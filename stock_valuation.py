"""
Оценка справедливой стоимости акции.

Источники данных:
  - MOEX ISS API  (цена, дивиденды)        — для российских тикеров
  - Claude API    (парсинг PDF МСФО отчёта) — EPS, BVPS, ROE, g
  - yfinance                                — для зарубежных тикеров
  - Ручной ввод                             — fallback

Модели оценки:  DDM (Gordon), P/E (компаративный), RIV (остаточный доход)

Настройка API ключа — создай файл .env рядом со скриптом:
    ANTHROPIC_API_KEY=sk-ant-api03-...

Запуск:
    python stock_valuation.py SBER report.pdf    # с PDF отчётом ← рекомендуется
    python stock_valuation.py SBER               # без PDF (ручной ввод)
    python stock_valuation.py AAPL report.pdf    # зарубежные акции
"""

import os
import sys
import json
import re
import base64
import requests
import re
from datetime import datetime, timedelta
try:
    from bs4 import BeautifulSoup
    BS4_OK = True
except ImportError:
    BS4_OK = False


# ────────────────────────────────────────────────────────────────
#  .env loader (без внешних зависимостей)
# ────────────────────────────────────────────────────────────────

def _load_dotenv():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip())

_load_dotenv()


# ────────────────────────────────────────────────────────────────
#  Утилиты
# ────────────────────────────────────────────────────────────────

def safe(val, default=0.0):
    try:
        v = float(val)
        return v if v == v else default
    except (TypeError, ValueError):
        return default


def ask_float(prompt, default):
    try:
        raw = input(f"  {prompt} [{default}]: ").strip()
        return float(raw) if raw else float(default)
    except ValueError:
        return float(default)


# ────────────────────────────────────────────────────────────────
#  Claude API — парсинг PDF МСФО отчёта
# ────────────────────────────────────────────────────────────────

IFRS_PROMPT = """
Ты — финансовый аналитик. Изучи приложенный МСФО/GAAP отчёт и извлеки данные.

Верни ТОЛЬКО валидный JSON без каких-либо пояснений, markdown или текста вокруг.

{
  "company_name": "Название компании",
  "report_year": 2024,
  "currency": "RUB",
  "shares_outstanding": 21586948000,
  "net_profit": 1508600000000,
  "equity": 6730000000000,
  "equity_begin": 5800000000000,
  "earnings_per_share": 33.6,
  "book_value_per_share": 215.0,
  "roe": 0.24,
  "dividend_per_share": 33.3,
  "dividend_growth_rate": 0.08,
  "revenue": 3800000000000,
  "revenue_prev": 3200000000000,
  "confidence": {
    "earnings_per_share": "high",
    "book_value_per_share": "high",
    "roe": "calculated",
    "dividend_per_share": "medium",
    "dividend_growth_rate": "calculated"
  },
  "notes": "Пояснения если что-то не найдено"
}

Правила:
- Все суммы в оригинальной валюте отчёта (НЕ пересчитывай)
- Если данных нет — ставь null
- EPS = Чистая прибыль / Кол-во акций (если нет готового значения)
- BVPS = Собственный капитал на конец периода / Кол-во акций
- ROE = Чистая прибыль / Собственный капитал на НАЧАЛО периода
- dividend_per_share — рекомендованный или объявленный дивиденд за отчётный год
- dividend_growth_rate — среднегодовой рост дивидендов за 3-5 лет (если есть история)
- confidence: "high"=нашёл напрямую, "calculated"=посчитал из других данных, "estimated"=оценка
"""


def extract_ifrs_from_pdf(pdf_path: str, api_key: str) -> dict:
    """Отправляет PDF в Claude API, получает структурированные финансовые данные."""
    size_mb = os.path.getsize(pdf_path) / 1024 / 1024
    print(f"\n🤖 Анализирую PDF через Claude API…")
    print(f"   Файл: {os.path.basename(pdf_path)}  ({size_mb:.1f} MB)")

    if size_mb > 32:
        print("  ⚠ Файл больше 32MB — Claude API может не принять. Попробуй сжать PDF.")

    with open(pdf_path, "rb") as f:
        pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")

    payload = {
        "model": "claude-opus-4-5",
        "max_tokens": 2000,
        "messages": [{
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_b64,
                    },
                },
                {"type": "text", "text": IFRS_PROMPT},
            ],
        }],
    }

    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers,
            json=payload,
            timeout=180,
        )
        r.raise_for_status()
    except requests.HTTPError as e:
        print(f"  ❌ Ошибка API: {e}\n  {r.text[:300]}")
        raise

    raw = r.json()["content"][0]["text"].strip()
    # Убираем ```json ... ``` если вдруг обернул
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    data = json.loads(raw)
    cur = data.get("currency", "")

    print(f"\n  ✅ Отчёт: {data.get('company_name','?')}  ({data.get('report_year','?')})")
    print(f"  {'─'*45}")

    rows = [
        ("EPS — прибыль на акцию",        "earnings_per_share",    cur, False),
        ("BVPS — балансовая стоимость",    "book_value_per_share",  cur, False),
        ("ROE — рентабельность капитала",  "roe",                   "",  True),
        ("Дивиденд на акцию (D0)",         "dividend_per_share",    cur, False),
        ("Рост дивидендов g",              "dividend_growth_rate",  "",  True),
    ]
    for label, key, unit, is_pct in rows:
        val  = data.get(key)
        conf = data.get("confidence", {}).get(key, "")
        if val is not None:
            disp = f"{val*100:.1f}%" if is_pct else f"{val:,.2f} {unit}"
            print(f"  {label}: {disp}  [{conf}]")
        else:
            print(f"  {label}: ⚠ не найдено")

    if data.get("notes"):
        print(f"\n  📝 {data['notes']}")

    return data



# ────────────────────────────────────────────────────────────────
#  Smart-lab.ru — фундаментальные данные (бесплатно, без API)
# ────────────────────────────────────────────────────────────────

SMARTLAB_HDR = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9",
}


def _sfloat(s: str) -> float:
    """Безопасный парсинг числа из строки (пробелы, запятые, минусы)."""
    s = re.sub(r'\s+', '', str(s or '')).replace(',', '.').replace('−','−'.replace('−','-')).replace('–','-')
    s = re.sub(r'[^0-9.\-]', '', s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def fetch_smartlab(ticker: str, verbose: bool = True) -> dict:
    """
    Парсит smart-lab.ru/q/TICKER/f/y/ — годовая МСФО отчётность.
    Возвращает dict с ключами: eps, bvps, roe, g  (или пустой dict при ошибке).
    verbose=False — без вывода в консоль (для фоновых запросов).
    """
    if not BS4_OK:
        if verbose:
            print("  ⚠ beautifulsoup4 не установлен: pip install beautifulsoup4")
        return {}

    ticker_clean = ticker.upper().replace(".ME", "")
    url = f"https://smart-lab.ru/q/{ticker_clean}/f/y/"
    if verbose:
        print(f"  📊 Загружаю фундаментальные данные с smart-lab.ru…")

    try:
        r = requests.get(url, headers=SMARTLAB_HDR, timeout=20)
        r.raise_for_status()
    except Exception as e:
        if verbose:
            print(f"  ⚠ smart-lab недоступен: {e}")
        return {}

    soup = BeautifulSoup(r.text, "html.parser")

    # Ищем таблицу — на smart-lab она имеет класс trades-table или simple-little-table
    table = None
    for t in soup.find_all("table"):
        text = t.get_text()
        if re.search(r'EPS|Прибыль на акцию|ROE|Выручка', text, re.I):
            table = t
            break

    if not table:
        if verbose:
            print("  ⚠ Таблица с данными не найдена на smart-lab")
        return {}

    # Собираем строки: {название_строки: [значение_по_годам]}
    rows_data = {}
    for row in table.find_all("tr"):
        cells = row.find_all(["td", "th"])
        if len(cells) < 2:
            continue
        name = cells[0].get_text(strip=True)
        vals = [cells[i].get_text(strip=True) for i in range(1, len(cells))]
        if name:
            rows_data[name] = vals

    def last_val(patterns):
        """Найти значение по паттерну — берём последнее непустое число."""
        for name, vals in rows_data.items():
            for pat in patterns:
                if re.search(pat, name, re.I):
                    for v in reversed(vals):
                        if v and v not in ('-', '—', 'None', ''):
                            result = _sfloat(v)
                            if result != 0.0:
                                return result
        return 0.0

    eps  = last_val([r'\bEPS\b', r'Прибыль на акцию'])
    bvps = last_val([r'\bBV\b', r'Балансовая стоимость.*акц', r'Book value per share'])
    roe  = last_val([r'\bROE\b'])
    if roe > 1:      # smart-lab отдаёт в процентах (24.5, а не 0.245)
        roe /= 100

    # Рост дивидендов g — считаем из истории дивидендов на странице
    g = 0.08
    div_vals = []
    for name, vals in rows_data.items():
        if re.search(r'Дивиденд|DPS|Dividend', name, re.I):
            nums = [_sfloat(v) for v in vals if v and v not in ('-','—')]
            nums = [x for x in nums if x > 0]
            if len(nums) >= 2:
                div_vals = nums
            break

    if len(div_vals) >= 3:
        try:
            g_calc = (div_vals[-1] / div_vals[-3]) ** 0.5 - 1
            if 0 < g_calc < 0.5:
                g = round(g_calc, 3)
        except (ZeroDivisionError, ValueError):
            pass

    if eps == 0 and bvps == 0:
        if verbose:
            print("  ⚠ Данные с smart-lab не распознаны (возможно сайт изменил структуру)")
        return {}

    if verbose:
        print(f"  ✅ Smart-lab: EPS={eps:.2f}₽  BVPS={bvps:.2f}₽  ROE={roe*100:.1f}%  g≈{g*100:.1f}%")
    return {"eps": eps, "bvps": bvps, "roe": roe, "g": g}

# ────────────────────────────────────────────────────────────────
#  Авто-получение ставок из открытых источников
# ────────────────────────────────────────────────────────────────

def fetch_cbr_key_rate() -> tuple:
    """
    Актуальная ключевая ставка ЦБ РФ через официальный XML-API cbr.ru.
    Возвращает (дата_строкой, ставка_в_долях), например ('2025-02-14', 0.21).
    При ошибке возвращает (None, None).
    """
    import xml.etree.ElementTree as ET
    url = "https://www.cbr.ru/scripts/XML_keyrate.asp"
    try:
        r = requests.get(url, timeout=10,
                         headers={"User-Agent": "StockValuator/1.0"})
        r.raise_for_status()
        root = ET.fromstring(r.content)
        rates = []
        for kr in root.findall("KR"):
            date_s = kr.get("Date", "")
            rate_s = kr.get("Rate", "0").replace(",", ".")
            try:
                rates.append((date_s, float(rate_s) / 100))
            except ValueError:
                pass
        if rates:
            rates.sort(key=lambda x: x[0])
            return rates[-1]   # самая свежая запись
    except Exception as e:
        print(f"  ⚠ Ошибка загрузки ставки ЦБ РФ: {e}")
    return (None, None)


def fetch_moex_market_return(years: int = 5) -> tuple:
    """
    Историческая доходность российского рынка акций.
    Использует MCFTR (индекс полной доходности MOEX — с учётом дивидендов).
    Fallback: IMOEX (ценовой индекс).
    Возвращает (cagr_float, описание) или (None, None) при ошибке.
    """
    till  = datetime.now().strftime("%Y-%m-%d")
    start = (datetime.now() - timedelta(days=years * 365 + 60)).strftime("%Y-%m-%d")

    for ticker in ("MCFTR", "IMOEX"):
        url = (f"https://iss.moex.com/iss/engines/stock/markets/index/"
               f"boards/SNDX/securities/{ticker}/candles.json"
               f"?from={start}&till={till}&interval=24"
               f"&iss.meta=off&iss.json=extended")
        try:
            r = requests.get(url, timeout=15,
                             headers={"User-Agent": "StockValuator/1.0"})
            r.raise_for_status()
            closes, dates = [], []
            for block in r.json():
                if not isinstance(block, dict):
                    continue
                for row in block.get("candles", []):
                    try:
                        closes.append(float(row["close"]))
                        dates.append(row["begin"][:10])
                    except Exception:
                        pass
            if len(closes) >= 50:
                actual_years = (datetime.strptime(dates[-1], "%Y-%m-%d") -
                                datetime.strptime(dates[0],  "%Y-%m-%d")).days / 365.25
                cagr = (closes[-1] / closes[0]) ** (1 / actual_years) - 1
                desc = (f"{ticker}  {dates[0]}→{dates[-1]}  "
                        f"{closes[0]:.0f}→{closes[-1]:.0f}  "
                        f"CAGR {cagr*100:.1f}%")
                return (round(cagr, 4), desc)
        except Exception as e:
            print(f"  ⚠ {ticker}: {e}")

    return (None, None)


# ────────────────────────────────────────────────────────────────
#  MOEX ISS API
# ────────────────────────────────────────────────────────────────

MOEX_BASE = "https://iss.moex.com/iss"
MOEX_HDR  = {"Accept": "application/json", "User-Agent": "StockValuator/1.0"}


def moex_get(url, params=None):
    p = dict(params or {})
    p["iss.json"] = "extended"
    p["iss.meta"] = "off"
    r = requests.get(url, params=p, headers=MOEX_HDR, timeout=15)
    r.raise_for_status()
    return r.json()


def moex_price(ticker: str) -> float:
    url = f"{MOEX_BASE}/engines/stock/markets/shares/boards/TQBR/securities/{ticker}.json"
    try:
        data = moex_get(url)
        for block in data:
            if not isinstance(block, dict):
                continue
            for key in ("marketdata", "securities"):
                for row in block.get(key, []):
                    for field in ("LAST", "CLOSE", "PREVPRICE", "WAPRICE"):
                        v = safe(row.get(field))
                        if v > 0:
                            return v
    except Exception as e:
        print(f"  ⚠ moex_price: {e}")
    return 0.0


def moex_dividends(ticker: str) -> list:
    url = f"{MOEX_BASE}/securities/{ticker}/dividends.json"
    try:
        data = moex_get(url)
        divs = []
        for block in data:
            if not isinstance(block, dict):
                continue
            for row in block.get("dividends", []):
                v = safe(row.get("value") or row.get("VALUE"))
                d = row.get("registryclosedate") or row.get("REGISTRYCLOSEDATE") or ""
                if v > 0:
                    divs.append((d, v))
        divs.sort(key=lambda x: x[0])
        return divs
    except Exception as e:
        print(f"  ⚠ moex_dividends: {e}")
        return []


def moex_name(ticker: str) -> str:
    try:
        data = moex_get(f"{MOEX_BASE}/securities/{ticker}.json")
        for block in data:
            if not isinstance(block, dict):
                continue
            info = {r.get("name",""): r.get("value","")
                    for r in block.get("description", []) if isinstance(r, dict)}
            name = info.get("NAME") or info.get("SHORTNAME")
            if name:
                return name
    except Exception:
        pass
    return ticker


# ────────────────────────────────────────────────────────────────
#  Основная функция сборки данных (MOEX + PDF или ручной ввод)
# ────────────────────────────────────────────────────────────────

def fetch_moex(ticker: str, pdf_path: str = None) -> dict:
    ticker = ticker.upper().replace(".ME", "")
    print(f"\n📡 Загружаю рыночные данные с MOEX ISS для {ticker}…")

    price = moex_price(ticker)
    divs  = moex_dividends(ticker)
    name  = moex_name(ticker)

    if price == 0:
        print(f"  ⚠ Цена не найдена через API")
        price = ask_float(f"Текущая цена акции {ticker}, ₽", 0.0)

    # D0 из истории дивидендов MOEX
    if divs:
        cutoff = datetime.now().replace(year=datetime.now().year - 1).strftime("%Y-%m-%d")
        recent = [v for d, v in divs if d >= cutoff]
        d0_moex = sum(recent) if recent else sum(v for _, v in divs[-4:])
        print(f"  ✅ Цена (P_0):               {price:.2f} ₽")
        print(f"  ✅ Дивиденд D0 (год):        {d0_moex:.2f} ₽  "
              f"(последние выплаты: {[round(v,2) for _,v in divs[-4:]]})")
    else:
        d0_moex = 0.0
        print(f"  ✅ Цена (P_0): {price:.2f} ₽")
        print(f"  ⚠ История дивидендов MOEX не найдена")

    # ── Фундаментальные данные: PDF или ручной ввод ──────────────
    eps = bvps = roe = g = beta = None
    d0 = d0_moex

    if pdf_path:
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            print("\n  ❌ ANTHROPIC_API_KEY не найден!")
            print("     Создай файл .env рядом со скриптом:")
            print("     ANTHROPIC_API_KEY=sk-ant-api03-...")
            print("\n  ⚙️  Переходим к ручному вводу…")
        else:
            try:
                ifrs = extract_ifrs_from_pdf(pdf_path, api_key)
                eps  = safe(ifrs.get("earnings_per_share"))
                bvps = safe(ifrs.get("book_value_per_share"))
                roe  = safe(ifrs.get("roe"))
                g    = safe(ifrs.get("dividend_growth_rate"), 0.08)
                # Дивиденд из PDF приоритетнее если найден
                d0_pdf = safe(ifrs.get("dividend_per_share"))
                if d0_pdf > 0:
                    d0 = d0_pdf
                    print(f"\n  ℹ Используем дивиденд из PDF: {d0:.2f} ₽"
                          f" (MOEX: {d0_moex:.2f} ₽)")
            except Exception as e:
                print(f"\n  ❌ Ошибка парсинга PDF: {e}")
                print("  ⚙️  Переходим к ручному вводу…")

    # Ручной ввод для отсутствующих / уточнение автоматических
    # Если нет данных из PDF — пробуем smart-lab
    if not pdf_path and (eps is None or eps == 0):
        sl = fetch_smartlab(ticker)
        if sl:
            eps  = sl.get("eps",  eps)
            bvps = sl.get("bvps", bvps)
            roe  = sl.get("roe",  roe)
            g    = sl.get("g",    g)

    # Если BVPS не найден автоматически — он критичен для RIV, предупреждаем
    if not bvps or bvps == 0:
        print(f"\n  ⚠ BVPS не найден автоматически (нужен для RIV модели).")
        print(f"    Возьми из последнего отчёта: Собственный капитал / Кол-во акций")

    # Показываем отраслевой P/E который будет использован
    sector_pe_val = get_sector_pe(ticker)
    print(f"\n  ℹ Отраслевой P/E для {ticker}: {sector_pe_val} (используется в P/E методе)")
    print(f"  ⚙️  Проверь данные (Enter = оставить найденное, 0 = не найдено):")
    eps  = ask_float("EPS — прибыль на акцию, ₽",               eps  if eps  else 0.0)
    bvps = ask_float("BVPS — балансовая стоимость на акцию, ₽", bvps if bvps else 0.0)
    g    = ask_float("g — ожидаемый рост дивидендов (0.10=10%)", g   if g    else 0.08)
    roe  = ask_float("ROE — рентабельность капитала (0.20=20%)", roe if roe  else 0.15)
    beta = ask_float("β бета акции (1.0 = среднерыночная)",      1.0)

    if d0_moex > 0 and d0 != d0_moex:
        confirm = input(f"\n  Дивиденд: PDF={d0:.2f} vs MOEX={d0_moex:.2f}. "
                       f"Какой использовать? [pdf/moex, Enter=pdf]: ").strip().lower()
        if confirm == "moex":
            d0 = d0_moex

    pe = (price / eps) if eps > 0 else 0.0

    # Ставки дисконтирования для РФ
    r_f    = 0.16   # Ключевая ставка ЦБ РФ
    r_m    = 0.22   # Историческая доходность IMOEX
    r_capm = r_f + beta * (r_m - r_f)
    d1     = d0 * (1 + g)
    r_ddm  = (d1 / price + g) if price > 0 and d1 > 0 else r_capm
    r_ddm  = max(min(r_ddm, 0.60), r_f)
    r_avg  = (r_capm + r_ddm) / 2

    T     = 5
    ri    = (roe - r_avg) * bvps
    pv_ri = sum(ri / (1 + r_avg) ** t for t in range(1, T + 1)) if r_avg else 0

    return dict(
        ticker=ticker, name=name,
        price=price,
        d0=d0, d1=d1, g=g, k=r_avg,
        eps=eps, pe=pe,
        bvps=bvps, pv_ri=pv_ri, roe=roe,
        beta=beta, r_f=r_f, r_m=r_m,
        r_capm=r_capm, r_ddm=r_ddm, r_avg=r_avg,
        currency="₽",
    )


# ────────────────────────────────────────────────────────────────
#  yfinance (зарубежные акции)
# ────────────────────────────────────────────────────────────────

def fetch_yfinance(ticker: str, pdf_path: str = None) -> dict:
    try:
        import yfinance as yf
        import pandas as pd
    except ImportError:
        print("  ⚠ yfinance не установлен: pip install yfinance")
        sys.exit(1)

    print(f"\n📡 Загружаю данные через yfinance для {ticker}…")
    t    = yf.Ticker(ticker)
    info = t.info

    hist  = t.history(period="5d")
    price = float(hist["Close"].iloc[-1]) if not hist.empty else safe(info.get("currentPrice"))

    div_hist = t.dividends
    if not div_hist.empty:
        cutoff = div_hist.index[-1] - pd.DateOffset(years=1)
        d0     = float(div_hist[div_hist.index >= cutoff].sum())
        d0     = d0 if d0 > 0 else float(div_hist.iloc[-1])
    else:
        d0 = safe(info.get("trailingAnnualDividendRate") or info.get("dividendRate"))

    eps  = safe(info.get("trailingEps") or info.get("forwardEps"))
    bvps = safe(info.get("bookValue"))
    roe  = safe(info.get("returnOnEquity"), 0.15)
    g    = safe(info.get("earningsGrowth") or info.get("revenueGrowth"), 0.05)
    g    = g if 0 < g < 0.50 else 0.05
    pe   = safe(info.get("trailingPE") or info.get("forwardPE"))
    pe   = pe if pe > 0 else (price / eps if eps > 0 else 0)
    beta = safe(info.get("beta"), 1.0)

    # PDF override для зарубежных
    if pdf_path:
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if api_key:
            try:
                ifrs = extract_ifrs_from_pdf(pdf_path, api_key)
                eps  = safe(ifrs.get("earnings_per_share"),  eps)
                bvps = safe(ifrs.get("book_value_per_share"), bvps)
                roe  = safe(ifrs.get("roe"),                 roe)
                g    = safe(ifrs.get("dividend_growth_rate"), g)
                d0_pdf = safe(ifrs.get("dividend_per_share"))
                if d0_pdf > 0:
                    d0 = d0_pdf
            except Exception as e:
                print(f"  ❌ PDF parsing error: {e}")

    r_f    = 0.045
    r_m    = 0.10
    r_capm = r_f + beta * (r_m - r_f)
    d1     = d0 * (1 + g)
    r_ddm  = (d1 / price + g) if price > 0 and d1 > 0 else r_capm
    r_ddm  = max(min(r_ddm, 0.40), r_f)
    r_avg  = (r_capm + r_ddm) / 2

    T     = 5
    ri    = (roe - r_avg) * bvps
    pv_ri = sum(ri / (1 + r_avg) ** t for t in range(1, T + 1)) if r_avg else 0

    return dict(
        ticker=ticker, name=info.get("longName", ticker),
        price=price,
        d0=d0, d1=d1, g=g, k=r_avg,
        eps=eps, pe=pe,
        bvps=bvps, pv_ri=pv_ri, roe=roe,
        beta=beta, r_f=r_f, r_m=r_m,
        r_capm=r_capm, r_ddm=r_ddm, r_avg=r_avg,
        currency="$",
    )


# ────────────────────────────────────────────────────────────────
#  Модели оценки
# ────────────────────────────────────────────────────────────────

# Отраслевые P/E для российского рынка — базовые значения (fallback)
SECTOR_PE = {
    "Нефть и газ":       5.5,
    "Банки":             5.0,
    "Металлы":           6.0,
    "Ритейл":            8.0,
    "Телеком":           7.0,
    "Технологии":       15.0,
    "Электроэнергетика": 5.0,
    "Удобрения":         6.5,
    "Транспорт":         7.0,
    "Девелопмент":       4.5,
    "default":           6.0,
}

# Маппинг тикеров на сектор (расширяй по необходимости)
TICKER_SECTOR = {
    "SBER": "Банки",   "VTBR": "Банки",   "TCSG": "Банки",   "BSPB": "Банки",
    "LKOH": "Нефть и газ", "ROSN": "Нефть и газ", "SNGS": "Нефть и газ",
    "NVTK": "Нефть и газ", "TATN": "Нефть и газ", "BANE": "Нефть и газ",
    "GMKN": "Металлы", "NLMK": "Металлы", "CHMF": "Металлы",
    "MAGN": "Металлы", "ALRS": "Металлы", "POLY": "Металлы",
    "MGNT": "Ритейл",  "FIVE": "Ритейл",  "OZON": "Технологии",
    "YNDX": "Технологии", "T":  "Технологии",
    "MTSS": "Телеком", "RTKM": "Телеком",
    "FEES": "Электроэнергетика", "HYDR": "Электроэнергетика",
    "PHOR": "Удобрения", "AKRN": "Удобрения",
    "FLOT": "Транспорт", "AFLT": "Транспорт",
    "SMLT": "Девелопмент", "PIKK": "Девелопмент",
}

# Репрезентативные тикеры на сектор для расчёта реального P/E
SECTOR_REPS = {
    "Банки":             ["SBER", "VTBR"],
    "Нефть и газ":       ["LKOH", "TATN"],
    "Металлы":           ["GMKN", "NLMK"],
    "Ритейл":            ["MGNT", "FIVE"],
    "Телеком":           ["MTSS", "RTKM"],
    "Технологии":        ["YNDX", "OZON"],
    "Электроэнергетика": ["FEES", "HYDR"],
    "Удобрения":         ["PHOR", "AKRN"],
    "Транспорт":         ["FLOT", "AFLT"],
    "Девелопмент":       ["SMLT"],
}

# Живые P/E: обновляются из фонового потока (app.py → _refresh_sector_pe)
_LIVE_SECTOR_PE: dict = {}


def get_sector_pe(ticker: str) -> float:
    """Возвращает отраслевой P/E: сначала из живых данных, иначе базовый fallback."""
    sector = TICKER_SECTOR.get(ticker.upper(), "default")
    live   = _LIVE_SECTOR_PE.get(sector)
    if live and live > 0:
        return live
    return SECTOR_PE.get(sector, SECTOR_PE["default"])


def fetch_sector_pe_live() -> dict:
    """
    Для каждого сектора берём репрезентативные тикеры,
    параллельно запрашиваем цену (MOEX) + EPS (SmartLab),
    вычисляем медианный P/E.
    Возвращает dict {sector: float}.  Тихий режим — без вывода в консоль.
    """
    import statistics
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _ticker_pe(ticker: str):
        try:
            price = moex_price(ticker)
            if price <= 0:
                return None
            sl = fetch_smartlab(ticker, verbose=False)
            eps = sl.get("eps", 0.0) if sl else 0.0
            if eps <= 0:
                return None
            pe = price / eps
            return pe if 1.0 < pe < 150.0 else None   # sanity
        except Exception:
            return None

    result = {}
    # Все тикеры разом — максимум параллельности
    all_tickers = [(sector, t) for sector, tickers in SECTOR_REPS.items()
                               for t in tickers]
    pe_by_sector: dict = {s: [] for s in SECTOR_REPS}

    with ThreadPoolExecutor(max_workers=min(len(all_tickers), 10)) as ex:
        futures = {ex.submit(_ticker_pe, t): (sector, t)
                   for sector, t in all_tickers}
        for f in as_completed(futures):
            sector, _ = futures[f]
            pe = f.result()
            if pe is not None:
                pe_by_sector[sector].append(pe)

    for sector, pes in pe_by_sector.items():
        if pes:
            result[sector] = round(statistics.median(pes), 1)
        else:
            result[sector] = SECTOR_PE.get(sector, SECTOR_PE["default"])

    all_vals = list(result.values())
    result["default"] = round(statistics.median(all_vals), 1) if all_vals else SECTOR_PE["default"]
    return result


def ddm_price(d: dict) -> float:
    """Gordon Growth Model: P = D1 / (k - g_ddm).
    g_ddm = min(g, 20%) — ограничение роста в DDM.
    Защита: если (k-g_ddm) < 5% — модель ненадёжна, возвращаем 0."""
    k  = d["k"]
    d0 = d.get("d0", 0.0)
    if d0 <= 0:
        return 0.0
    g_ddm  = min(d["g"], 0.20)
    d1_ddm = d0 * (1 + g_ddm)
    spread = k - g_ddm
    if spread < 0.05:
        return 0.0
    return d1_ddm / spread


def pe_price(d: dict) -> float:
    """Метод сравнений: P = EPS × отраслевой P/E (не собственный)."""
    eps = d["eps"]
    if eps <= 0:
        return 0.0
    sector_pe = get_sector_pe(d["ticker"])
    return eps * sector_pe


def riv_price(d: dict) -> float:
    """Residual Income Valuation: P = BVPS + PV остаточных доходов.""";
    bvps = d["bvps"]
    if bvps <= 0:
        return 0.0
    return bvps + d["pv_ri"]


def dcf_price(d: dict) -> float:
    """Упрощённый DCF через нормализованный FCF.

    FCF = EPS × payout_ratio  (дивиденды / прибыль, но не >1).
    Прогнозируем FCF на T лет с ростом g_proj = min(g, 20%),
    затем терминальная стоимость с g_term = min(g, 4%).
    Минимальный спред k − g_term ≥ 5% — защита от взрыва TV.

    P = Σ FCF*(1+g_proj)^t / (1+k)^t  +  TV / (1+k)^T
    TV = FCF_T * (1 + g_term) / (k − g_term)
    """
    eps, g, k = d["eps"], d["g"], d["k"]
    if eps <= 0 or k <= 0:
        return 0.0

    # Fix 1: FCF = EPS × payout_ratio (а не чистый EPS)
    d0 = d.get("d0", 0.0)
    if d0 > 0 and eps > 0:
        payout = min(d0 / eps, 1.0)
    else:
        payout = 0.5          # по умолчанию 50%
    fcf = eps * payout

    # Fix 3: прогнозный рост не выше 20% — защита от гиперроста
    g_proj = min(g, 0.20)

    T      = 7
    # Fix 2: терминальный рост ≤ 4%, а спред (k − g_term) ≥ 5%
    g_term = min(g, 0.04)
    if k - g_term < 0.05:
        g_term = k - 0.05
    if g_term < 0 or k <= g_term:
        return 0.0

    pv_fcf = sum(fcf * (1 + g_proj) ** t / (1 + k) ** t for t in range(1, T + 1))
    fcf_T  = fcf * (1 + g_proj) ** T
    tv     = fcf_T * (1 + g_term) / (k - g_term)
    pv_tv  = tv / (1 + k) ** T
    return pv_fcf + pv_tv


# ────────────────────────────────────────────────────────────────
#  Excel
# ────────────────────────────────────────────────────────────────

def write_excel(d: dict, wb):
    pass  # детальные листы отключены — только сводка


def update_summary_sheet(wb, d: dict):
    """Единственный лист portfolio.xlsx — строка на каждый тикер."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN_DARK  = "1D5C3A"
    HEADER_BG   = "2E7D52"
    ALT         = "F2F9F5"
    GREEN_LIGHT = "C6EFCE"
    RED_LIGHT   = "FFCCCC"

    def brd():
        s = Side(style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    SHEET = "Portfolio"

    # ── Считываем существующие строки ────────────────────────────
    existing = {}   # ticker -> row list
    if SHEET in wb.sheetnames:
        ws_old = wb[SHEET]
        for row in ws_old.iter_rows(min_row=3, values_only=True):
            if row and row[0]:
                existing[str(row[0])] = list(row)
        del wb[SHEET]

    # ── Обновляем текущий тикер ───────────────────────────────────
    ddm = ddm_price(d); cmp = pe_price(d); riv = riv_price(d)
    models = [v for v in [ddm, cmp, riv] if v > 0]
    avg    = sum(models) / len(models) if models else 0
    upside = round((avg / d["price"] - 1) * 100, 1) if d["price"] and avg else 0

    dcf = dcf_price(d)
    models = [v for v in [ddm, cmp, riv, dcf] if v > 0]
    avg    = sum(models) / len(models) if models else 0
    upside = round((avg / d["price"] - 1) * 100, 1) if d["price"] and avg else 0

    existing[d["ticker"]] = [
        d["ticker"],
        d["name"],
        d["currency"],
        round(d["price"],        2),   # Текущая цена
        round(ddm,               2),   # DDM
        round(cmp,               2),   # P/E отраслевой
        round(riv,               2),   # RIV
        round(dcf,               2),   # DCF
        round(avg,               2),   # Среднее (справедливая цена)
        upside,                        # Потенциал %
        round(d["eps"],          2),   # EPS
        round(d["price"] / d["eps"] if d["eps"] > 0 else 0, 1),  # Текущий P/E
        round(get_sector_pe(d["ticker"]), 1),  # Отраслевой P/E
        round(d["bvps"],         2),   # BVPS
        round(d["roe"]*100,      1),   # ROE %
        round(d["g"]*100,        1),   # g %
        round(d["k"]*100,        2),   # k %
        round(d["d0"],           2),   # D0 дивиденд
        round(d["d1"],           2),   # D1 прогноз
        datetime.now().strftime("%d.%m.%Y"),
    ]

    # ── Создаём лист ─────────────────────────────────────────────
    ws = wb.create_sheet(SHEET, 0)

    COLS = [
        ("Тикер",       8),
        ("Компания",    28),
        ("Вал.",        5),
        ("Цена",        10),
        ("DDM",         11),
        ("P/E отрасль", 11),
        ("RIV",         11),
        ("DCF",         11),
        ("Справедл.",   11),
        ("Потенц.%",    10),
        ("EPS",          8),
        ("Тек.P/E",      8),
        ("Отр.P/E",      8),
        ("BVPS",         9),
        ("ROE%",         7),
        ("g%",           6),
        ("k%",           6),
        ("D0",           8),
        ("D1",           8),
        ("Обновлено",   12),
    ]

    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Заголовок
    ncols = len(COLS)
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"Portfolio  |  обновлено {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    c.font  = Font(bold=True, color="FFFFFF", size=12)
    c.fill  = PatternFill("solid", fgColor=GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Шапка колонок
    for ci, (label, _) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd()
    ws.row_dimensions[2].height = 22

    # Строки данных — сортировка по потенциалу (убывание)
    sorted_rows = sorted(
        existing.values(),
        key=lambda r: r[8] if isinstance(r[8], (int, float)) else -999,
        reverse=True
    )

    for ri, row_vals in enumerate(sorted_rows, 3):
        bg = ALT if ri % 2 == 1 else "FFFFFF"
        upside_val = row_vals[8] if len(row_vals) > 8 else 0

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = brd()
            c.font   = Font(size=10)
            c.alignment = Alignment(
                horizontal="left" if ci == 2 else "center",
                vertical="center"
            )

            # Колонка "Потенциал%" — цветная
            if ci == 9:
                if isinstance(val, (int, float)) and val > 0:
                    c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
                    c.font = Font(bold=True, color="1D5C3A", size=10)
                elif isinstance(val, (int, float)) and val < 0:
                    c.fill = PatternFill("solid", fgColor=RED_LIGHT)
                    c.font = Font(bold=True, color="C00000", size=10)
                else:
                    c.fill = PatternFill("solid", fgColor=bg)
            # Колонка "Справедл." — зелёная
            elif ci == 8:
                c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
                c.font = Font(bold=True, color="1D5C3A", size=10)
            else:
                c.fill = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[ri].height = 18


def main():
    args = sys.argv[1:]

    ticker   = None
    pdf_path = None

    for arg in args:
        if arg.lower().endswith(".pdf"):
            if not os.path.exists(arg):
                print(f"❌ PDF файл не найден: {arg}")
                sys.exit(1)
            pdf_path = arg
        else:
            ticker = arg.strip().upper()

    if not ticker:
        ticker = input("Введите тикер (SBER, LKOH, AAPL…): ").strip().upper()

    if not pdf_path:
        pdf_input = input("Путь к PDF отчёту (Enter — пропустить): ").strip()
        if pdf_input:
            if not os.path.exists(pdf_input):
                print(f"❌ Файл не найден: {pdf_input}")
                sys.exit(1)
            pdf_path = pdf_input

    is_moex = not ("." in ticker) or ticker.endswith(".ME")

    if is_moex:
        data = fetch_moex(ticker, pdf_path)
    else:
        data = fetch_yfinance(ticker, pdf_path)

    ddm = ddm_price(data);  cmp = pe_price(data)
    riv = riv_price(data);  dcf = dcf_price(data)
    models = [v for v in [ddm, cmp, riv, dcf] if v > 0]
    avg    = sum(models) / len(models) if models else 0
    cur    = data["currency"]
    sp     = get_sector_pe(data["ticker"])

    print(f"\n{'─'*50}")
    print(f"  💰 ОЦЕНКА: {data['ticker']}  ({data['name']})")
    print(f"{'─'*50}")
    ddm_note = "" if ddm > 0 else "  ← (k-g) < 5%, пропущено"
    print(f"  DDM (Gordon):        {ddm:>10.2f} {cur}{ddm_note}")
    print(f"  P/E отраслевой ({sp}): {cmp:>8.2f} {cur}")
    riv_note = "" if riv > 0 else "  ← BVPS не задан"
    print(f"  RIV (остат. доход):  {riv:>10.2f} {cur}{riv_note}")
    print(f"  DCF (7 лет):         {dcf:>10.2f} {cur}")
    print(f"  {'─'*38}")
    print(f"  Среднее (из {len(models)} моделей): {avg:>8.2f} {cur}")
    print(f"  Текущая цена:        {data['price']:>10.2f} {cur}")
    diff  = (avg / data["price"] - 1) * 100 if data["price"] and avg else 0
    arrow = "📈" if diff > 0 else "📉"
    print(f"  {arrow} Потенциал:       {diff:>+9.1f}%")
    print(f"{'─'*50}\n")

    script_dir   = os.path.dirname(os.path.abspath(__file__))
    portfolio_file = os.path.join(script_dir, "portfolio.xlsx")

    try:
        import openpyxl
    except ImportError:
        print("  ⚠ openpyxl не установлен: pip install openpyxl")
        return

    # Открываем существующий файл или создаём новый
    if os.path.exists(portfolio_file):
        wb = openpyxl.load_workbook(portfolio_file)
        print(f"\n📂 Обновляю существующий файл: portfolio.xlsx")
    else:
        wb = openpyxl.Workbook()
        # Удалим пустой дефолтный лист
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        print(f"\n📂 Создаю новый файл: portfolio.xlsx")

    write_excel(data, wb)

    # Обновляем/создаём сводный лист с итогами
    update_summary_sheet(wb, data)

    wb.save(portfolio_file)
    ticker_clean = ticker.replace(".ME", "")
    print(f"✅ Сохранено: portfolio.xlsx  (листы: {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
