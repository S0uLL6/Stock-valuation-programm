"""
Stock Valuation Desktop App
pip install customtkinter matplotlib requests beautifulsoup4 openpyxl
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog
import threading, sys, os, requests, math, json
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.patches as mpatches
import numpy as np
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import stock_valuation as sv
from stock_valuation import (
    ddm_price, pe_price, riv_price, dcf_price,
    get_sector_pe, update_summary_sheet,
    moex_dividends, moex_price, moex_name, fetch_smartlab,
    fetch_cbr_key_rate, fetch_moex_market_return,
    fetch_sector_pe_live, SECTOR_PE,
    safe, _load_dotenv,
)
_load_dotenv()

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Конфиг (ставки дисконтирования) ────────────────────────────
_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
_CONFIG_DEFAULTS = {"r_f": 0.16, "r_m": 0.22}
CONFIG = dict(_CONFIG_DEFAULTS)

def _load_config():
    global CONFIG
    if os.path.exists(_CONFIG_PATH):
        try:
            with open(_CONFIG_PATH, "r") as f:
                CONFIG.update(json.load(f))
        except Exception:
            pass

def _save_config():
    with open(_CONFIG_PATH, "w") as f:
        json.dump(CONFIG, f, indent=2)

_load_config()

# ── Портфель (персистентность) ──────────────────────────────────
_PORTFOLIO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "portfolio_state.json")

def _save_portfolio(portfolio: dict):
    try:
        serializable = {}
        for ticker, d in portfolio.items():
            serializable[ticker] = {k: v for k, v in d.items()
                                    if isinstance(v, (str, int, float, bool, type(None)))}
        with open(_PORTFOLIO_PATH, "w", encoding="utf-8") as f:
            json.dump(serializable, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"⚠ Ошибка сохранения портфеля: {e}")

def _load_portfolio() -> dict:
    if os.path.exists(_PORTFOLIO_PATH):
        try:
            with open(_PORTFOLIO_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

# ── Цвета ──────────────────────────────────────────────────────
BG      = "#0D1117"
CARD    = "#161B22"
CARD2   = "#1C2128"
BORDER  = "#30363D"
GREEN   = "#3FB950"
GREEN_D = "#0D2818"
RED     = "#F85149"
RED_D   = "#2D1017"
GOLD    = "#E3B341"
GOLD_D  = "#2D2008"
TEXT    = "#E6EDF3"
MUTED   = "#8B949E"
ACCENT  = "#58A6FF"
PURPLE  = "#BC8CFF"

MPL = {
    "figure.facecolor": CARD, "axes.facecolor": CARD2,
    "axes.edgecolor": BORDER, "axes.labelcolor": MUTED,
    "xtick.color": MUTED, "ytick.color": MUTED,
    "grid.color": BORDER, "grid.alpha": 0.5,
    "text.color": TEXT, "lines.linewidth": 2,
}
plt.rcParams.update(MPL)
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

# ── Хелперы ────────────────────────────────────────────────────
def lbl(parent, text, size=12, weight="normal", color=TEXT, **kw):
    return ctk.CTkLabel(parent, text=text,
        font=ctk.CTkFont(family="SF Pro Display", size=size, weight=weight),
        text_color=color, **kw)

def inp(parent, width=140, ph="", **kw):
    return ctk.CTkEntry(parent, width=width, placeholder_text=ph,
        fg_color=CARD2, border_color=BORDER, text_color=TEXT,
        placeholder_text_color=MUTED, font=ctk.CTkFont(size=13), **kw)

def btn(parent, text, cmd, color=ACCENT, width=120, height=34, **kw):
    return ctk.CTkButton(parent, text=text, command=cmd,
        fg_color=color, hover_color=BORDER, text_color=TEXT,
        font=ctk.CTkFont(size=13, weight="bold"),
        corner_radius=8, width=width, height=height, **kw)

class Card(ctk.CTkFrame):
    def __init__(self, p, **kw):
        super().__init__(p, fg_color=CARD, corner_radius=12,
                         border_width=1, border_color=BORDER, **kw)

class Div(ctk.CTkFrame):
    def __init__(self, p, **kw):
        super().__init__(p, fg_color=BORDER, height=1, **kw)


# ═══════════════════════════════════════════════════════════════
#  Всплывающая подсказка для карточек моделей
# ═══════════════════════════════════════════════════════════════
MODEL_INFO = {
    "ddm": ("DDM — Gordon Growth Model",
            "P = D₁ / (k − g)\n\n"
            "D₁ — ожидаемый дивиденд\n"
            "k  — ставка дисконтирования\n"
            "g  — темп роста дивидендов\n\n"
            "Работает только если k − g ≥ 5%.\n"
            "Подходит для зрелых дивидендных компаний."),
    "pe":  ("P/E — Отраслевой метод",
            "P = EPS × P/E_отрасли\n\n"
            "EPS — прибыль на акцию\n"
            "P/E_отрасли — медианный мультипликатор\n"
            "              по сектору компании\n\n"
            "Показывает цену относительно\n"
            "среднерыночной оценки аналогов."),
    "riv": ("RIV — Residual Income Valuation",
            "P = BVPS + Σ (ROE − k) × BVPS / (1+k)ᵗ\n\n"
            "BVPS — балансовая стоимость на акцию\n"
            "ROE  — рентабельность капитала\n"
            "k    — ставка дисконтирования\n\n"
            "Учитывает создание стоимости\n"
            "сверх требуемой доходности."),
    "dcf": ("DCF — Дисконтированные потоки",
            "P = Σ EPS×(1+g)ᵗ/(1+k)ᵗ + TV/(1+k)ᵀ\n\n"
            "T  — горизонт прогноза (7 лет)\n"
            "TV — терминальная стоимость\n"
            "g_term = min(g, 4%)\n\n"
            "Наиболее полная модель: учитывает\n"
            "рост прибыли и долгосрочную стоимость."),
}

class TooltipWindow(tk.Toplevel):
    def __init__(self, parent, title, body):
        super().__init__(parent)
        self.overrideredirect(True)
        self.configure(bg=CARD2)
        self.attributes("-topmost", True)

        frame = tk.Frame(self, bg=CARD2, padx=16, pady=12)
        frame.pack()

        tk.Label(frame, text=title, bg=CARD2, fg=ACCENT,
                 font=("SF Pro Display", 13, "bold")).pack(anchor="w")
        tk.Frame(frame, bg=BORDER, height=1).pack(fill="x", pady=8)
        tk.Label(frame, text=body, bg=CARD2, fg=TEXT,
                 font=("SF Pro Mono", 11), justify="left").pack(anchor="w")

        # Позиционируем рядом с курсором
        x = parent.winfo_pointerx() + 12
        y = parent.winfo_pointery() + 12
        self.geometry(f"+{x}+{y}")

        self.bind("<Button-1>", lambda e: self.destroy())
        self.after(6000, self.destroy)


# ═══════════════════════════════════════════════════════════════
#  Страница 1 — Оценка
# ═══════════════════════════════════════════════════════════════
class ValuationPage(ctk.CTkFrame):
    def __init__(self, parent, app, **kw):
        super().__init__(parent, fg_color=BG, **kw)
        self.app = app
        self.data = None
        self.pdf_path = None
        self.portfolio = _load_portfolio()
        self._loaded_price = 0
        self._loaded_name  = ""
        self._loaded_d0    = 0
        self._tooltip_win  = None
        self._spinner_angle = 0
        self._spinning = False
        self._auto_on   = False
        self._auto_job  = None
        self._build()
        if self.portfolio:
            self.after(200, self._refresh_table)
            threading.Thread(target=self._refresh_portfolio_prices,
                             daemon=True).start()

    def _refresh_portfolio_prices(self):
        """Обновляет цены всех тикеров в портфеле (фон, при старте)."""
        for ticker, d in list(self.portfolio.items()):
            try:
                new_price = moex_price(ticker)
                if new_price > 0:
                    d["price"] = new_price
                    if d.get("avg", 0) > 0:
                        d["upside"] = round((d["avg"] / new_price - 1) * 100, 1)
            except Exception:
                pass
        _save_portfolio(self.portfolio)
        self.after(0, self._refresh_table)

    def _build(self):
        self.grid_columnconfigure(0, weight=0, minsize=300)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self._build_left()
        self._build_results()
        self._build_portfolio()

    # ── Левая панель ───────────────────────────────────────────
    def _build_left(self):
        p = Card(self)
        p.grid(row=0, column=0, rowspan=2, padx=(16,8), pady=16, sticky="nsew")
        p.grid_columnconfigure(0, weight=1)

        lbl(p, "Тикер акции", size=11, color=MUTED).grid(
            row=0, padx=16, pady=(16,4), sticky="w")

        row0 = ctk.CTkFrame(p, fg_color="transparent")
        row0.grid(row=1, padx=16, sticky="ew")
        row0.grid_columnconfigure(0, weight=1)

        self.ticker_e = inp(row0, ph="SBER, LKOH, AAPL…")
        self.ticker_e.grid(row=0, column=0, sticky="ew", padx=(0,8))
        self.ticker_e.bind("<Return>", lambda e: self._load_data())

        # Кнопка с spinner
        self.load_btn = btn(row0, "Загрузить", self._load_data,
                            color=ACCENT, width=100)
        self.load_btn.grid(row=0, column=1)

        # Spinner canvas
        self.spin_canvas = tk.Canvas(p, width=20, height=20,
                                     bg=CARD, highlightthickness=0)
        self.spin_canvas.grid(row=2, padx=16, pady=(6,0), sticky="w")

        # Авто-обновление цены
        auto_row = ctk.CTkFrame(p, fg_color="transparent")
        auto_row.grid(row=3, padx=16, pady=(4, 0), sticky="w")
        self._auto_btn = ctk.CTkButton(
            auto_row, text="⟳ Авто: выкл", width=110, height=26,
            fg_color=CARD2, hover_color=BORDER,
            font=ctk.CTkFont(size=11), command=self._toggle_auto)
        self._auto_btn.pack(side="left", padx=(0, 6))
        self._interval_var = tk.StringVar(value="5м")
        ctk.CTkOptionMenu(
            auto_row, values=["1м", "5м", "10м"],
            variable=self._interval_var,
            width=60, height=26,
            fg_color=CARD2, button_color=CARD2,
            button_hover_color=BORDER,
            font=ctk.CTkFont(size=11),
            dropdown_fg_color=CARD,
        ).pack(side="left")

        self.pdf_lbl = lbl(p, "PDF: не выбран", size=10, color=MUTED)
        self.pdf_lbl.grid(row=4, padx=16, pady=(4,0), sticky="w")
        btn(p, "Прикрепить PDF", self._pick_pdf,
            color=CARD2, width=200).grid(row=5, padx=16, pady=(4,0), sticky="w")

        Div(p).grid(row=6, padx=16, pady=10, sticky="ew")
        lbl(p, "Фундаментальные данные", size=11, color=MUTED).grid(
            row=7, padx=16, pady=(0,8), sticky="w")

        self.fields = {}
        for i, (key, lb, ph) in enumerate([
            ("eps",  "EPS, ₽",           "Прибыль на акцию"),
            ("bvps", "BVPS, ₽",          "Балансовая стоимость / акция"),
            ("roe",  "ROE (0.20 = 20%)", "Рентабельность капитала"),
            ("g",    "g (0.08 = 8%)",    "Рост дивидендов"),
            ("beta", "β (бета)",          "Бета акции"),
        ]):
            r = 8 + i * 2
            lbl(p, lb, size=11, color=MUTED).grid(
                row=r, padx=16, pady=(4,0), sticky="w")
            e = inp(p, ph=ph)
            e.grid(row=r+1, padx=16, pady=(2,0), sticky="ew")
            self.fields[key] = e

        Div(p).grid(row=18, padx=16, pady=10, sticky="ew")

        calc_row = ctk.CTkFrame(p, fg_color="transparent")
        calc_row.grid(row=19, padx=16, pady=(0,4), sticky="ew")
        calc_row.grid_columnconfigure(0, weight=1)
        self.calc_btn = btn(calc_row, "Рассчитать", self._calculate,
                            color="#238636", height=40)
        self.calc_btn.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        btn(calc_row, "✕ Сбросить", self._reset_fields,
            color=CARD2, width=90, height=40).grid(row=0, column=1)

        # Блок добавления в портфель
        add_frame = ctk.CTkFrame(p, fg_color=CARD2, corner_radius=8)
        add_frame.grid(row=20, padx=16, pady=(4,0), sticky="ew")
        add_frame.grid_columnconfigure(0, weight=1)

        lbl(add_frame, "Доля в портфеле, %", size=11,
            color=MUTED).grid(row=0, padx=12, pady=(8,2), sticky="w")
        self.weight_e = inp(add_frame, ph="например 10")
        self.weight_e.grid(row=1, padx=12, pady=(0,4), sticky="ew")
        btn(add_frame, "В портфель", self._add_to_portfolio,
            color="#1C4A2A", width=200).grid(
            row=2, padx=12, pady=(0,8), sticky="ew")

        btn(p, "Анализ чувствительности", self._show_sensitivity,
            color=CARD2, width=200).grid(
            row=21, padx=16, pady=(4,0), sticky="ew")

        btn(p, "Сценарный анализ", self._show_scenarios,
            color=CARD2, width=200).grid(
            row=22, padx=16, pady=(4,0), sticky="ew")

        btn(p, "Экспорт Excel", self._export_excel,
            color=CARD2, width=200).grid(
            row=23, padx=16, pady=(8,2), sticky="ew")

        btn(p, "Экспорт CSV", self._export_csv,
            color=CARD2, width=200).grid(
            row=24, padx=16, pady=(2,4), sticky="ew")

        self.status_lbl = lbl(p, "", size=11, color=MUTED)
        self.status_lbl.grid(row=25, padx=16, pady=(4,12), sticky="w")

    # ── Карточки результатов ───────────────────────────────────
    def _build_results(self):
        f = ctk.CTkFrame(self, fg_color="transparent")
        f.grid(row=0, column=1, padx=(8,16), pady=(16,8), sticky="ew")
        f.grid_columnconfigure((0,1,2,3,4), weight=1)

        self.model_cards = {}
        models = [
            ("ddm", "DDM",  "Gordon Growth Model", ACCENT),
            ("pe",  "P/E",  "Отраслевой метод",    PURPLE),
            ("riv", "RIV",  "Остаточный доход",    GOLD),
            ("dcf", "DCF",  "Дисконт. потоки",     "#79C0FF"),
        ]
        for col, (key, name, sub, color) in enumerate(models):
            card = Card(f)
            card.grid(row=0, column=col, padx=4, sticky="nsew")
            card.grid_columnconfigure(0, weight=1)
            card.configure(cursor="hand2")

            lbl(card, name, size=13, weight="bold", color=color).grid(
                row=0, padx=14, pady=(12,0), sticky="w")
            lbl(card, sub, size=10, color=MUTED).grid(
                row=1, padx=14, pady=(0,4), sticky="w")
            vl = lbl(card, "—", size=22, weight="bold", color=TEXT)
            vl.grid(row=2, padx=14, pady=(0,12), sticky="w")

            # Клик — показываем формулу
            for widget in (card, vl):
                widget.bind("<Button-1>",
                    lambda e, k=key: self._show_model_info(k))
            lbl(card, "нажми для формулы", size=9,
                color=MUTED).grid(row=3, padx=14, pady=(0,10), sticky="w")

            self.model_cards[key] = vl

        # P/E мини-карточка (row=1, span all columns)
        pe_card = Card(f)
        pe_card.grid(row=1, column=0, columnspan=5, padx=4, pady=(6,0), sticky="ew")
        pe_card.grid_columnconfigure((0,1,2,3), weight=1)
        lbl(pe_card, "P/E текущий", size=11, color=MUTED).grid(
            row=0, column=0, padx=14, pady=(8,2), sticky="w")
        lbl(pe_card, "P/E отраслевой", size=11, color=MUTED).grid(
            row=0, column=1, padx=14, pady=(8,2), sticky="w")
        lbl(pe_card, "Оценка", size=11, color=MUTED).grid(
            row=0, column=2, padx=14, pady=(8,2), sticky="w")
        self._pe_cur_lbl    = lbl(pe_card, "—", size=18, weight="bold")
        self._pe_cur_lbl.grid(row=1, column=0, padx=14, pady=(0,8), sticky="w")
        self._pe_sec_lbl    = lbl(pe_card, "—", size=18, weight="bold", color=MUTED)
        self._pe_sec_lbl.grid(row=1, column=1, padx=14, pady=(0,8), sticky="w")
        self._pe_verdict_lbl = lbl(pe_card, "—", size=13, color=MUTED)
        self._pe_verdict_lbl.grid(row=1, column=2, padx=14, pady=(0,8), sticky="w")

        # Итог
        self.fair_card = Card(f)
        self.fair_card.grid(row=0, column=4, padx=(8,0), sticky="nsew")
        self.fair_card.grid_columnconfigure(0, weight=1)

        lbl(self.fair_card, "Справедливая цена", size=13,
            weight="bold", color=GREEN).grid(
            row=0, padx=14, pady=(12,0), sticky="w")
        lbl(self.fair_card, "Среднее моделей", size=10,
            color=MUTED).grid(row=1, padx=14, pady=(0,4), sticky="w")
        self.fair_val = lbl(self.fair_card, "—", size=26,
                             weight="bold", color=GREEN)
        self.fair_val.grid(row=2, padx=14, pady=(0,4), sticky="w")
        self.upside_lbl = lbl(self.fair_card, "", size=14, color=MUTED)
        self.upside_lbl.grid(row=3, padx=14, pady=(0,10), sticky="w")
        self.models_used = lbl(self.fair_card, "", size=10, color=MUTED)
        self.models_used.grid(row=4, padx=14, pady=(0,12), sticky="w")

    # ── Таблица портфеля ───────────────────────────────────────
    def _build_portfolio(self):
        f = ctk.CTkFrame(self, fg_color="transparent")
        f.grid(row=1, column=1, padx=(8,16), pady=(0,16), sticky="nsew")
        f.grid_columnconfigure(0, weight=1)
        f.grid_rowconfigure(2, weight=1)

        hdr_row = ctk.CTkFrame(f, fg_color="transparent")
        hdr_row.grid(row=0, sticky="ew", pady=(0, 8))
        hdr_row.grid_columnconfigure(0, weight=1)
        lbl(hdr_row, "Портфель", size=13, weight="bold").grid(
            row=0, column=0, sticky="w")
        self._weight_warn = lbl(hdr_row, "", size=11, color=GOLD)
        self._weight_warn.grid(row=0, column=1, sticky="e", padx=(0, 8))
        btn(hdr_row, "Нормировать", self._normalize_weights,
            color=CARD2, width=110, height=26).grid(
            row=0, column=2, sticky="e", padx=(0, 6))
        btn(hdr_row, "Очистить", self._clear_portfolio,
            color=RED_D, width=90, height=26).grid(
            row=0, column=3, sticky="e")

        COLS   = ["Тикер","Компания","Доля %","Цена","DDM","P/E","RIV","DCF","Справедл.","Потенц.%",""]
        WIDTHS = [65, 180, 65, 75, 85, 85, 85, 85, 90, 85, 36]

        hdr = Card(f)
        hdr.grid(row=1, sticky="ew")
        for ci, (c, w) in enumerate(zip(COLS, WIDTHS)):
            lbl(hdr, c, size=11, color=MUTED).grid(
                row=0, column=ci, padx=6, pady=6, sticky="w")
            hdr.grid_columnconfigure(ci, minsize=w)

        self.tbl = ctk.CTkScrollableFrame(
            f, fg_color=CARD, corner_radius=0,
            border_width=1, border_color=BORDER)
        self.tbl.grid(row=2, sticky="nsew")
        for ci, w in enumerate(WIDTHS):
            self.tbl.grid_columnconfigure(ci, minsize=w)

    def _refresh_table(self):
        for w in self.tbl.winfo_children():
            w.destroy()

        sorted_items = sorted(
            self.portfolio.items(),
            key=lambda x: x[1].get("upside", 0), reverse=True)

        for ri, (ticker, d) in enumerate(sorted_items):
            bg = CARD if ri % 2 == 0 else CARD2
            up = d.get("upside", 0)
            uc = GREEN if up > 0 else (RED if up < 0 else MUTED)
            wt = d.get("weight", "—")
            wt_str = f"{wt}%" if isinstance(wt, (int,float)) else "—"

            vals = [
                ticker,
                d.get("name","")[:22],
                wt_str,
                f"{d['price']:.2f}",
                f"{d['ddm']:.2f}" if d['ddm'] else "—",
                f"{d['cmp']:.2f}" if d['cmp'] else "—",
                f"{d['riv']:.2f}" if d['riv'] else "—",
                f"{d['dcf']:.2f}" if d['dcf'] else "—",
                f"{d['avg']:.2f}" if d['avg'] else "—",
                f"{up:+.1f}%",
                "✕",
            ]
            for ci, val in enumerate(vals):
                color = uc if ci == 9 else (RED if ci == 10 else TEXT)
                if ci == 2:  # Доля — редактируемое поле
                    e = ctk.CTkEntry(self.tbl, width=55, height=24,
                        fg_color=CARD2, border_color=BORDER,
                        text_color=TEXT, font=ctk.CTkFont(size=12))
                    e.insert(0, wt_str.replace("%",""))
                    e.grid(row=ri, column=ci, padx=4, pady=2, sticky="w")
                    e.bind("<Return>",
                        lambda ev, t=ticker, en=e: self._update_weight(t, en))
                    e.bind("<FocusOut>",
                        lambda ev, t=ticker, en=e: self._update_weight(t, en))
                elif ci == 10:  # Кнопка удаления
                    b = ctk.CTkButton(
                        self.tbl, text="✕", width=28, height=24,
                        fg_color="transparent", hover_color=RED_D,
                        text_color=MUTED, font=ctk.CTkFont(size=11),
                        command=lambda t=ticker: self._remove(t))
                    b.grid(row=ri, column=ci, padx=4, pady=2)
                else:
                    cl = ctk.CTkLabel(
                        self.tbl, text=val,
                        font=ctk.CTkFont(size=12),
                        text_color=color, fg_color=bg,
                        anchor="w", padx=6)
                    cl.grid(row=ri, column=ci, sticky="ew", ipady=4)

        if sorted_items:
            self._append_total_row(len(sorted_items))

    def _calc_portfolio_stats(self):
        """Возвращает (weighted_upside, total_weight) по портфелю."""
        weighted_sum = 0.0
        weight_sum   = 0.0
        for d in self.portfolio.values():
            w  = d.get("weight")
            up = d.get("upside", 0)
            if isinstance(w, (int, float)) and w > 0:
                weighted_sum += up * w
                weight_sum   += w
        w_upside = weighted_sum / weight_sum if weight_sum else 0
        return w_upside, weight_sum

    def _append_total_row(self, ri):
        """Добавляет строку «Итого» после всех тикеров."""
        w_upside, weight_sum = self._calc_portfolio_stats()
        diff = round(weight_sum - 100, 1)
        if weight_sum == 0:
            self._weight_warn.configure(text="")
        elif abs(diff) < 0.1:
            self._weight_warn.configure(text="✓ 100%", text_color=GREEN)
        else:
            sign = "+" if diff > 0 else ""
            self._weight_warn.configure(
                text=f"⚠ {weight_sum:.1f}% ({sign}{diff:.1f}%)",
                text_color=GOLD)
        up_color = GREEN if w_upside > 0 else (RED if w_upside < 0 else MUTED)
        up_str   = f"{w_upside:+.1f}%" if weight_sum else "—"
        wt_str   = f"{weight_sum:.1f}%"

        # Разделитель
        sep = ctk.CTkFrame(self.tbl, fg_color=BORDER, height=1)
        sep.grid(row=ri, column=0, columnspan=11, sticky="ew", pady=(4, 2))

        total_vals = ["", "Итого", wt_str, "", "", "", "", "", "", up_str, ""]
        for ci, val in enumerate(total_vals):
            color = up_color if ci == 9 else (GOLD if ci == 1 else MUTED)
            cl = ctk.CTkLabel(
                self.tbl, text=val,
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=color, fg_color=CARD,
                anchor="w", padx=6)
            cl.grid(row=ri + 1, column=ci, sticky="ew", ipady=4)

    def _update_weight(self, ticker, entry_widget):
        try:
            val = float(entry_widget.get().replace("%","").strip())
            self.portfolio[ticker]["weight"] = val
            _save_portfolio(self.portfolio)
        except ValueError:
            pass

    def _remove(self, ticker):
        self.portfolio.pop(ticker, None)
        _save_portfolio(self.portfolio)
        self._refresh_table()

    def _normalize_weights(self):
        """Приводит все веса к 100%, пропорционально."""
        _, weight_sum = self._calc_portfolio_stats()
        if not weight_sum:
            return
        for d in self.portfolio.values():
            w = d.get("weight")
            if isinstance(w, (int, float)) and w > 0:
                d["weight"] = round(w / weight_sum * 100, 1)
        _save_portfolio(self.portfolio)
        self._refresh_table()

    def _clear_portfolio(self):
        from tkinter import messagebox
        if not self.portfolio:
            return
        if messagebox.askyesno("Очистить портфель",
                               "Удалить все тикеры из портфеля?"):
            self.portfolio.clear()
            _save_portfolio(self.portfolio)
            self._refresh_table()

    # ── Spinner ────────────────────────────────────────────────
    def _start_spinner(self):
        self._spinning = True
        self._spinner_angle = 0
        self._tick_spinner()

    def _tick_spinner(self):
        if not self._spinning:
            self.spin_canvas.delete("all")
            return
        c = self.spin_canvas
        c.delete("all")
        cx, cy, r = 10, 10, 7
        for i in range(8):
            a = math.radians(self._spinner_angle + i * 45)
            alpha = int(255 * (i + 1) / 8)
            color = f"#{alpha:02x}{alpha:02x}{alpha:02x}"
            x1 = cx + (r-2) * math.cos(a)
            y1 = cy + (r-2) * math.sin(a)
            x2 = cx + r * math.cos(a)
            y2 = cy + r * math.sin(a)
            c.create_line(x1,y1,x2,y2, fill=color, width=2)
        self._spinner_angle = (self._spinner_angle + 15) % 360
        self.after(50, self._tick_spinner)

    def _stop_spinner(self):
        self._spinning = False

    # ── Подсказка с формулой ───────────────────────────────────
    def _show_model_info(self, key):
        if self._tooltip_win:
            try:
                self._tooltip_win.destroy()
            except Exception:
                pass
        title, body = MODEL_INFO[key]
        self._tooltip_win = TooltipWindow(self, title, body)

    # ── Статус ─────────────────────────────────────────────────
    def _status(self, msg, color=MUTED):
        self.status_lbl.configure(text=msg, text_color=color)

    def _toggle_auto(self):
        self._auto_on = not self._auto_on
        if self._auto_on:
            self._auto_btn.configure(text="⟳ Авто: вкл", fg_color=ACCENT)
            self._do_auto_update()
        else:
            self._auto_btn.configure(text="⟳ Авто: выкл", fg_color=CARD2)
            if self._auto_job:
                self.after_cancel(self._auto_job)
                self._auto_job = None

    def _do_auto_update(self):
        if not self._auto_on:
            return
        ticker = self.ticker_e.get().strip().upper().replace(".ME", "")
        if ticker:
            def worker():
                try:
                    price = moex_price(ticker)
                    if price > 0:
                        self.after(0, lambda: self._apply_auto_price(ticker, price))
                except Exception:
                    pass
            threading.Thread(target=worker, daemon=True).start()
        interval_map = {"1м": 60_000, "5м": 300_000, "10м": 600_000}
        ms = interval_map.get(self._interval_var.get(), 300_000)
        self._auto_job = self.after(ms, self._do_auto_update)

    def _flash_price(self, color, n=4):
        """Мигание border карточки справедливой цены при обновлении."""
        if n <= 0:
            return
        current = self.fair_card.cget("border_color")
        next_color = color if current != color else BORDER
        self.fair_card.configure(border_color=next_color)
        self.after(180, lambda: self._flash_price(color, n - 1))

    def _apply_auto_price(self, ticker, price):
        prev_price = self._loaded_price
        self._loaded_price = price
        ts = datetime.now().strftime("%H:%M:%S")
        # Пересчитываем upside если есть данные расчёта
        if self.data and self.data.get("ticker") == ticker:
            avg = self.data.get("avg", 0)
            upside = (avg / price - 1) * 100 if price and avg else 0
            self.data["price"] = price
            self.data["upside"] = upside
            if upside > 0:
                self.upside_lbl.configure(
                    text=f"▲ +{upside:.1f}% потенциал", text_color=GREEN)
                self.fair_card.configure(border_color=GREEN)
            elif upside < 0:
                self.upside_lbl.configure(
                    text=f"▼ {upside:.1f}% переоценена", text_color=RED)
                self.fair_card.configure(border_color=RED)
        if prev_price and abs(price - prev_price) > 0.001:
            flash_color = GREEN if price >= prev_price else RED
            self._flash_price(flash_color)
        self._status(f"⟳ Обновлено: {ts}  Цена: {price:.2f} ₽", MUTED)

    def _pick_pdf(self):
        p = filedialog.askopenfilename(
            filetypes=[("PDF", "*.pdf")], title="Выбери МСФО отчёт")
        if p:
            self.pdf_path = p
            self.pdf_lbl.configure(
                text=f"PDF: {os.path.basename(p)}", text_color=GREEN)

    def _load_data(self):
        ticker = self.ticker_e.get().strip().upper()
        if not ticker:
            self._status("Введи тикер", RED)
            return
        self.load_btn.configure(state="disabled", text="…")
        self._status("Загружаю…", GOLD)
        self._start_spinner()

        def worker():
            try:
                is_moex = not ("." in ticker) or ticker.endswith(".ME")
                tick = ticker.replace(".ME", "")
                if is_moex:
                    price = moex_price(tick)
                    name  = moex_name(tick)
                    sl    = fetch_smartlab(tick)
                    divs  = moex_dividends(tick)
                    if divs:
                        cutoff = datetime.now().replace(
                            year=datetime.now().year-1).strftime("%Y-%m-%d")
                        recent = [v for d,v in divs if d >= cutoff]
                        d0 = sum(recent) if recent else sum(v for _,v in divs[-4:])
                    else:
                        d0 = 0.0
                    self.after(0, lambda: self._fill({
                        "price":price,"name":name,"d0":d0,**sl}))
                else:
                    import yfinance as yf
                    t = yf.Ticker(ticker)
                    info = t.info
                    hist = t.history(period="5d")
                    price = float(hist["Close"].iloc[-1]) if not hist.empty else 0
                    self.after(0, lambda: self._fill({
                        "price":price,
                        "name":info.get("longName",ticker),
                        "eps": safe(info.get("trailingEps")),
                        "bvps":safe(info.get("bookValue")),
                        "roe": safe(info.get("returnOnEquity"),0.15),
                        "g":   0.05,
                        "beta":safe(info.get("beta"),1.0),
                    }))
            except Exception as e:
                self.after(0, lambda: self._status(f"Ошибка: {e}", RED))
            finally:
                self.after(0, self._stop_spinner)
                self.after(0, lambda: self.load_btn.configure(
                    state="normal", text="Загрузить"))

        threading.Thread(target=worker, daemon=True).start()

    def _fill(self, d):
        self._loaded_price = d.get("price", 0)
        self._loaded_name  = d.get("name", "")
        self._loaded_d0    = d.get("d0", 0)
        self._status(
            f"✓ {d.get('name','')}  Цена: {d.get('price',0):.2f} ₽", GREEN)
        for key in ("eps","bvps","roe","g","beta"):
            val = d.get(key)
            if val:
                self.fields[key].delete(0, "end")
                self.fields[key].insert(0, str(round(float(val), 4)))
        ticker = self.ticker_e.get().strip().upper().replace(".ME","")
        self.app.analytics_page.load_ticker(ticker)

    def _fval(self, key, default=0.0):
        try:
            return float(self.fields[key].get()) or default
        except (ValueError, TypeError):
            return default

    def _reset_fields(self):
        # 13.2 — Очистка полей ввода
        for e in self.fields.values():
            e.delete(0, "end")
        # 13.3 — Сброс карточек результатов
        for card in self.model_cards.values():
            card.configure(text="—")
        self.fair_val.configure(text="—")
        self.upside_lbl.configure(text="", text_color=MUTED)
        self.models_used.configure(text="")
        self.fair_card.configure(border_color=BORDER)
        self._pe_cur_lbl.configure(text="—", text_color=TEXT)
        self._pe_sec_lbl.configure(text="—", text_color=MUTED)
        self._pe_verdict_lbl.configure(text="—", text_color=MUTED)
        # 13.4 — Сброс PDF
        self.pdf_path = None
        self.pdf_lbl.configure(text="PDF: не выбран", text_color=MUTED)
        # 13.5 — Сброс загруженных данных
        self._loaded_price = 0
        self._loaded_d0    = 0
        self._loaded_name  = ""
        self.data = None
        self._status("Поля сброшены", MUTED)

    def _update_pe_card(self, cur_pe, sector_pe):
        if cur_pe > 0:
            self._pe_cur_lbl.configure(text=f"{cur_pe:.1f}")
        else:
            self._pe_cur_lbl.configure(text="—", text_color=MUTED)
            self._pe_sec_lbl.configure(text="—")
            self._pe_verdict_lbl.configure(text="нет EPS", text_color=MUTED)
            return
        self._pe_sec_lbl.configure(text=f"{sector_pe:.1f}", text_color=MUTED)
        if cur_pe < sector_pe:
            color   = GREEN
            verdict = f"▼ ниже отрасли на {sector_pe - cur_pe:.1f}x"
        elif cur_pe > sector_pe:
            color   = RED
            verdict = f"▲ выше отрасли на {cur_pe - sector_pe:.1f}x"
        else:
            color   = MUTED
            verdict = "= на уровне отрасли"
        self._pe_cur_lbl.configure(text_color=color)
        self._pe_verdict_lbl.configure(text=verdict, text_color=color)

    def _calculate(self):
        ticker = self.ticker_e.get().strip().upper().replace(".ME","")
        if not ticker or self._loaded_price == 0:
            self._status("Сначала загрузи данные", RED)
            return
        price = self._loaded_price
        d0    = self._loaded_d0
        name  = self._loaded_name

        eps  = self._fval("eps")
        bvps = self._fval("bvps")
        roe  = self._fval("roe", 0.15)
        g    = self._fval("g",   0.08)
        beta = self._fval("beta", 1.0)

        r_f    = CONFIG["r_f"]; r_m = CONFIG["r_m"]
        r_capm = r_f + beta * (r_m - r_f)
        d1     = d0 * (1 + g)
        r_ddm  = (d1/price + g) if price and d1 else r_capm
        r_ddm  = max(min(r_ddm, 0.60), r_f)
        r_avg  = (r_capm + r_ddm) / 2
        ri     = (roe - r_avg) * bvps
        pv_ri  = sum(ri/(1+r_avg)**t for t in range(1,6)) if r_avg else 0

        d = dict(ticker=ticker, name=name, price=price,
                 d0=d0, d1=d1, g=g, k=r_avg,
                 eps=eps, pe=(price/eps if eps>0 else 0),
                 bvps=bvps, pv_ri=pv_ri, roe=roe,
                 beta=beta, r_f=r_f, r_m=r_m,
                 r_capm=r_capm, r_ddm=r_ddm, r_avg=r_avg,
                 currency="₽")

        ddm = ddm_price(d); cmp = pe_price(d)
        riv = riv_price(d); dcf = dcf_price(d)
        models = [v for v in [ddm,cmp,riv,dcf] if v > 0]
        avg    = sum(models)/len(models) if models else 0
        upside = (avg/price - 1)*100 if price and avg else 0

        for key, val in [("ddm",ddm),("pe",cmp),("riv",riv),("dcf",dcf)]:
            self.model_cards[key].configure(
                text=f"{val:.2f} ₽" if val > 0 else "—")

        self.fair_val.configure(text=f"{avg:.2f} ₽" if avg else "—")
        self.models_used.configure(text=f"из {len(models)} моделей")

        if upside > 0:
            self.upside_lbl.configure(
                text=f"▲ +{upside:.1f}% потенциал", text_color=GREEN)
            self.fair_card.configure(border_color=GREEN)
        elif upside < 0:
            self.upside_lbl.configure(
                text=f"▼ {upside:.1f}% переоценена", text_color=RED)
            self.fair_card.configure(border_color=RED)

        cur_pe     = price / eps if eps > 0 else 0
        sector_pe  = get_sector_pe(ticker)
        self.data = {**d, "ddm":ddm,"cmp":cmp,"riv":riv,"dcf":dcf,
                     "avg":avg,"upside":upside,
                     "cur_pe": cur_pe, "sector_pe": sector_pe}
        self._update_pe_card(cur_pe, sector_pe)
        if g > 0.20:
            self._status(
                f"✓ k={r_avg*100:.1f}%  CAPM={r_capm*100:.1f}%  ⚠ g обрезано до 20%", GOLD)
        else:
            self._status(f"✓ k={r_avg*100:.1f}%  CAPM={r_capm*100:.1f}%", GREEN)

    def _show_sensitivity(self):
        if not self.data:
            self._status("Сначала рассчитай", RED)
            return

        d = self.data
        cur_g = d["g"]
        cur_k = d["k"]

        # Сетки значений
        g_vals = [round(0.03 + i * 0.025, 3) for i in range(7)]  # 3%–18%
        k_vals = [round(0.12 + i * 0.03,  3) for i in range(7)]  # 12%–33%

        # Расчёт средней справедливой цены для каждой комбинации g×k
        def fair(g, k):
            dd = dict(d, g=g, k=k,
                      d1=d["d0"] * (1 + g),
                      pv_ri=sum((d["roe"] - k) * d["bvps"] / (1 + k) ** t
                                for t in range(1, 6)))
            vals = [v for v in [ddm_price(dd), pe_price(dd),
                                riv_price(dd), dcf_price(dd)] if v > 0]
            return sum(vals) / len(vals) if vals else 0

        matrix = [[fair(g, k) for k in k_vals] for g in g_vals]
        price  = d["price"]

        # ── Popup окно ─────────────────────────────────────────
        win = tk.Toplevel(self)
        win.title(f"Анализ чувствительности — {d['ticker']}")
        win.configure(bg=CARD)
        win.resizable(False, False)

        tk.Label(win, text=f"Справедливая цена  {d['ticker']}  (текущая цена {price:.0f} ₽)",
                 bg=CARD, fg=TEXT, font=("SF Pro Display", 13, "bold")).pack(
                 padx=20, pady=(16, 4))
        tk.Label(win,
                 text="Строки — рост дивидендов g      Столбцы — ставка дисконтирования k",
                 bg=CARD, fg=MUTED, font=("SF Pro Display", 10)).pack(padx=20, pady=(0, 10))

        # ── Heatmap через matplotlib ────────────────────────────
        fig, ax = plt.subplots(figsize=(9, 5.5))
        fig.patch.set_facecolor(CARD)
        ax.set_facecolor(CARD2)

        import matplotlib.colors as mcolors
        upside_matrix = [[(fair(g, k) / price - 1) * 100
                          for k in k_vals] for g in g_vals]
        data_np = np.array(upside_matrix)

        # Цветовая карта: красный (−50%) → белый (0%) → зелёный (+50%)
        cmap = mcolors.LinearSegmentedColormap.from_list(
            "rg", ["#8B0000", "#2D1017", CARD2, GREEN_D, "#005000"], N=256)
        vmax = max(abs(data_np.min()), abs(data_np.max()), 20)
        im = ax.imshow(data_np, cmap=cmap, vmin=-vmax, vmax=vmax, aspect="auto")

        ax.set_xticks(range(len(k_vals)))
        ax.set_xticklabels([f"{k*100:.0f}%" for k in k_vals],
                           color=MUTED, fontsize=9)
        ax.set_yticks(range(len(g_vals)))
        ax.set_yticklabels([f"{g*100:.0f}%" for g in g_vals],
                           color=MUTED, fontsize=9)
        ax.set_xlabel("k — ставка дисконтирования", color=MUTED, fontsize=10)
        ax.set_ylabel("g — рост дивидендов", color=MUTED, fontsize=10)
        ax.tick_params(colors=MUTED)
        for sp in ax.spines.values():
            sp.set_color(BORDER)

        # Текст в ячейках
        for i, g in enumerate(g_vals):
            for j, k in enumerate(k_vals):
                val  = matrix[i][j]
                upsd = upside_matrix[i][j]
                clr  = "#FFFFFF" if abs(upsd) > 15 else TEXT
                ax.text(j, i, f"{val:.0f}\n{upsd:+.0f}%",
                        ha="center", va="center", fontsize=8,
                        color=clr, fontweight="bold")

        # Рамка текущих параметров
        try:
            ci = min(range(len(k_vals)), key=lambda j: abs(k_vals[j] - cur_k))
            ri = min(range(len(g_vals)), key=lambda i: abs(g_vals[i] - cur_g))
            rect = plt.Rectangle((ci - 0.5, ri - 0.5), 1, 1,
                                  fill=False, edgecolor=GOLD,
                                  linewidth=2.5, zorder=5)
            ax.add_patch(rect)
            ax.text(ci, ri - 0.55, "◆ текущие",
                    ha="center", va="top", fontsize=7,
                    color=GOLD, fontweight="bold")
        except Exception:
            pass

        cbar = fig.colorbar(im, ax=ax, fraction=0.03, pad=0.02)
        cbar.set_label("Потенциал, %", color=MUTED, fontsize=9)
        cbar.ax.yaxis.set_tick_params(color=MUTED)
        plt.setp(cbar.ax.yaxis.get_ticklabels(), color=MUTED)

        fig.tight_layout(pad=1.5)

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(padx=10, pady=(0, 12))
        canvas.draw()

        tk.Button(win, text="Закрыть", command=win.destroy,
                  bg=CARD2, fg=TEXT, relief="flat",
                  font=("SF Pro Display", 11),
                  padx=20, pady=6).pack(pady=(0, 14))

    def _show_scenarios(self):
        if not self.data:
            self._status("Сначала рассчитай", RED)
            return

        d = self.data
        price = d["price"]

        # Пресеты сценариев: (название, иконка, цвет, множитель g, множитель k)
        SCENARIOS = [
            ("Медведь", "🐻", RED,   0.5,  1.2),
            ("База",    "—",  MUTED, 1.0,  1.0),
            ("Бык",     "🐂", GREEN, 1.5,  0.85),
        ]

        win = tk.Toplevel(self)
        win.title(f"Сценарный анализ — {d['ticker']}")
        win.configure(bg=CARD)
        win.resizable(False, False)

        # Заголовок
        tk.Label(win, text=f"Сценарный анализ  {d['ticker']}",
                 bg=CARD, fg=TEXT,
                 font=("SF Pro Display", 14, "bold")).grid(
            row=0, column=0, columnspan=3, padx=20, pady=(16, 4))
        tk.Label(win, text=f"Текущая цена: {price:.2f} ₽",
                 bg=CARD, fg=MUTED,
                 font=("SF Pro Display", 11)).grid(
            row=1, column=0, columnspan=3, padx=20, pady=(0, 12))

        # Хранилище Entry и Label виджетов для каждого сценария
        scenario_widgets = []

        for col, (name, icon, color, g_mult, k_mult) in enumerate(SCENARIOS):
            g_preset = round(d["g"] * g_mult * 100, 1)
            k_preset = round(d["k"] * k_mult * 100, 1)

            frame = tk.Frame(win, bg=CARD2, relief="flat",
                             highlightbackground=color,
                             highlightthickness=2)
            frame.grid(row=2, column=col, padx=10, pady=(0, 10),
                       sticky="nsew", ipadx=10, ipady=10)

            # Название
            tk.Label(frame, text=f"{icon} {name}",
                     bg=CARD2, fg=color,
                     font=("SF Pro Display", 13, "bold")).pack(pady=(10, 6))

            # g
            tk.Label(frame, text="g, % (рост дивидендов)",
                     bg=CARD2, fg=MUTED,
                     font=("SF Pro Display", 10)).pack()
            g_var = tk.StringVar(value=str(g_preset))
            g_entry = tk.Entry(frame, textvariable=g_var, width=10,
                               bg=CARD, fg=TEXT, insertbackground=TEXT,
                               relief="flat", font=("SF Pro Display", 12),
                               justify="center")
            g_entry.pack(pady=(2, 8))

            # k
            tk.Label(frame, text="k, % (ставка дисконт.)",
                     bg=CARD2, fg=MUTED,
                     font=("SF Pro Display", 10)).pack()
            k_var = tk.StringVar(value=str(k_preset))
            k_entry = tk.Entry(frame, textvariable=k_var, width=10,
                               bg=CARD, fg=TEXT, insertbackground=TEXT,
                               relief="flat", font=("SF Pro Display", 12),
                               justify="center")
            k_entry.pack(pady=(2, 8))

            # Результаты
            tk.Label(frame, text="Справедливая цена",
                     bg=CARD2, fg=MUTED,
                     font=("SF Pro Display", 10)).pack(pady=(6, 0))
            fair_lbl = tk.Label(frame, text="—",
                                bg=CARD2, fg=TEXT,
                                font=("SF Pro Display", 14, "bold"))
            fair_lbl.pack()

            upside_lbl = tk.Label(frame, text="",
                                  bg=CARD2, fg=MUTED,
                                  font=("SF Pro Display", 12))
            upside_lbl.pack(pady=(2, 10))

            scenario_widgets.append((g_var, k_var, fair_lbl, upside_lbl))

        def _calc_scenarios():
            for g_var, k_var, fair_lbl, upside_lbl in scenario_widgets:
                try:
                    g_val = float(g_var.get().replace(",", ".")) / 100
                    k_val = float(k_var.get().replace(",", ".")) / 100
                except ValueError:
                    fair_lbl.configure(text="ошибка", fg=RED)
                    upside_lbl.configure(text="")
                    continue
                dd = dict(d, g=g_val, k=k_val,
                          d1=d["d0"] * (1 + g_val),
                          pv_ri=sum((d["roe"] - k_val) * d["bvps"] / (1 + k_val) ** t
                                    for t in range(1, 6)))
                vals = [v for v in [
                    ddm_price(dd), pe_price(dd), riv_price(dd), dcf_price(dd)
                ] if v > 0]
                avg = sum(vals) / len(vals) if vals else 0
                if avg <= 0:
                    fair_lbl.configure(text="—", fg=TEXT)
                    upside_lbl.configure(text="")
                    continue
                upside = (avg / price - 1) * 100
                fair_lbl.configure(text=f"{avg:.2f} ₽", fg=TEXT)
                up_color = GREEN if upside >= 0 else RED
                arrow = "▲" if upside >= 0 else "▼"
                upside_lbl.configure(
                    text=f"{arrow} {upside:+.1f}%", fg=up_color)

        # Кнопки
        btn_frame = tk.Frame(win, bg=CARD)
        btn_frame.grid(row=3, column=0, columnspan=3, pady=(0, 14))

        tk.Button(btn_frame, text="Рассчитать",
                  command=_calc_scenarios,
                  bg=ACCENT, fg=TEXT, relief="flat",
                  font=("SF Pro Display", 11),
                  padx=18, pady=6).pack(side="left", padx=8)

        tk.Button(btn_frame, text="Закрыть",
                  command=win.destroy,
                  bg=CARD2, fg=TEXT, relief="flat",
                  font=("SF Pro Display", 11),
                  padx=18, pady=6).pack(side="left", padx=8)

        # Сразу считаем при открытии
        _calc_scenarios()

    def _add_to_portfolio(self):
        if not self.data:
            self._status("Сначала рассчитай", RED)
            return
        try:
            w = float(self.weight_e.get().strip()) if self.weight_e.get().strip() else None
        except ValueError:
            w = None
        self.data["weight"] = w
        self.portfolio[self.data["ticker"]] = dict(self.data)
        _save_portfolio(self.portfolio)
        self._refresh_table()
        self._status(f"✓ {self.data['ticker']} добавлен в портфель", GREEN)

    def _export_excel(self):
        if not self.portfolio and not self.data:
            self._status("Нет данных", RED); return
        if not OPENPYXL_OK:
            self._status("pip install openpyxl", RED); return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile="portfolio.xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not path: return
        import openpyxl
        wb = openpyxl.load_workbook(path) if os.path.exists(path) \
             else openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        items = {**self.portfolio}
        if self.data:
            items[self.data["ticker"]] = self.data
        for d in items.values():
            update_summary_sheet(wb, d)
        wb.save(path)
        self._status(f"✓ {os.path.basename(path)}", GREEN)

    def _export_csv(self):
        import csv
        items = {**self.portfolio}
        if self.data:
            items[self.data["ticker"]] = self.data
        if not items:
            self._status("Нет данных", RED)
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", initialfile="portfolio.csv",
            filetypes=[("CSV", "*.csv")])
        if not path:
            return
        cols = [
            ("Тикер",           lambda d: d.get("ticker", "")),
            ("Компания",        lambda d: d.get("name", "")),
            ("Доля %",          lambda d: d.get("weight", "")),
            ("Цена",            lambda d: d.get("price", "")),
            ("DDM",             lambda d: d.get("ddm", "")),
            ("P/E",             lambda d: d.get("cmp", "")),
            ("RIV",             lambda d: d.get("riv", "")),
            ("DCF",             lambda d: d.get("dcf", "")),
            ("Справедл. цена",  lambda d: d.get("avg", "")),
            ("Потенциал %",     lambda d: d.get("upside", "")),
        ]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([c[0] for c in cols])
            for d in items.values():
                writer.writerow([c[1](d) for c in cols])
        self._status(f"✓ {os.path.basename(path)}", GREEN)


# ═══════════════════════════════════════════════════════════════
#  Страница 2 — Аналитика
# ═══════════════════════════════════════════════════════════════
PERIODS = {
    "1Д":   (1,    "60"),
    "1Н":   (7,    "60"),
    "1М":   (31,   "24"),
    "3М":   (90,   "24"),
    "1Г":   (365,  "24"),
    "3Г":   (1095, "7"),
    "5Л":   (1825, "7"),
    "Всё":  (0,    "7"),
}

class AnalyticsPage(ctk.CTkFrame):
    def __init__(self, parent, **kw):
        super().__init__(parent, fg_color=BG, **kw)
        self.current_ticker = None
        self._all_dates  = []
        self._all_closes = []
        self._divs_cache = []
        self._annot_price = None
        self._vline       = None
        self._pin_idx        = None
        self._pin_marker     = None
        self._price_span_patch = None
        self._cmp_ticker  = None
        self._cmp_dates   = []
        self._cmp_closes  = []
        self._norm_mode   = True   # True = нормализованный вид, False = абсолют
        self._build()

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=3)
        self.grid_rowconfigure(2, weight=2)

        # ── Шапка с периодами ──────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.grid(row=0, column=0, padx=16, pady=(14,6), sticky="ew")

        self.title_lbl = lbl(hdr, "Аналитика — загрузи тикер на странице «Оценка»",
                             size=15, weight="bold")
        self.title_lbl.pack(side="left")

        period_frame = ctk.CTkFrame(hdr, fg_color="transparent")
        period_frame.pack(side="right")
        lbl(period_frame, "Период:", size=11, color=MUTED).pack(
            side="left", padx=(0,8))

        self._period_btns = {}
        self._active_period = "1Г"
        for p in PERIODS:
            b = ctk.CTkButton(
                period_frame, text=p, width=42, height=28,
                fg_color=CARD2 if p != "1Г" else ACCENT,
                hover_color=BORDER, text_color=TEXT,
                font=ctk.CTkFont(size=12), corner_radius=6,
                command=lambda pp=p: self._set_period(pp))
            b.pack(side="left", padx=2)
            self._period_btns[p] = b

        # ── График цены ────────────────────────────────────────
        pc = Card(self)
        pc.grid(row=1, column=0, padx=16, pady=(0,6), sticky="nsew")
        pc.grid_columnconfigure(0, weight=1)
        pc.grid_rowconfigure(1, weight=1)

        price_hdr = ctk.CTkFrame(pc, fg_color="transparent")
        price_hdr.grid(row=0, padx=16, pady=(12,4), sticky="ew")
        lbl(price_hdr, "График цены", size=13,
            weight="bold", color=ACCENT).pack(side="left")
        self.price_info = lbl(price_hdr, "", size=12, color=MUTED)
        self.price_info.pack(side="right")

        # Поле сравнения тикеров
        self._cmp_entry = ctk.CTkEntry(
            price_hdr, placeholder_text="Сравнить с…",
            width=120, height=26, font=ctk.CTkFont(size=11))
        self._cmp_entry.pack(side="left", padx=(12, 4))
        self._cmp_entry.bind("<Return>", lambda e: self._load_compare())
        ctk.CTkButton(
            price_hdr, text="Сравнить", width=76, height=26,
            fg_color=CARD2, hover_color=BORDER,
            font=ctk.CTkFont(size=11),
            command=self._load_compare).pack(side="left", padx=(0, 4))
        self._cmp_clear_btn = ctk.CTkButton(
            price_hdr, text="✕", width=26, height=26,
            fg_color=CARD2, hover_color=RED,
            font=ctk.CTkFont(size=11),
            command=self._clear_compare)
        self._cmp_clear_btn.pack(side="left", padx=(0, 4))
        self._cmp_clear_btn.configure(state="disabled")
        self._norm_var = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            price_hdr, text="Норм", variable=self._norm_var,
            font=ctk.CTkFont(size=11), text_color=MUTED,
            fg_color=ACCENT, hover_color=BORDER,
            command=self._on_norm_toggle).pack(side="left", padx=(0, 12))

        # Чекбоксы индикаторов
        saved_ind = CONFIG.get("indicators", {})
        self._ind_vars = {}
        ind_cfg = [("SMA20","SMA 20"), ("SMA50","SMA 50"),
                   ("SMA200","SMA 200"), ("EMA20","EMA 20")]
        for key, label in ind_cfg:
            var = tk.BooleanVar(value=saved_ind.get(key, False))
            cb = ctk.CTkCheckBox(price_hdr, text=label, variable=var,
                                 font=ctk.CTkFont(size=11), text_color=MUTED,
                                 fg_color=ACCENT, hover_color=BORDER,
                                 command=self._on_indicator_toggle)
            cb.pack(side="right", padx=(0, 6))
            self._ind_vars[key] = var

        # Чекбокс «Объём»
        self._show_volume = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(price_hdr, text="Объём", variable=self._show_volume,
                        font=ctk.CTkFont(size=11), text_color=MUTED,
                        fg_color=ACCENT, hover_color=BORDER,
                        command=self._on_volume_toggle).pack(
                        side="right", padx=(0, 12))

        self.price_fig = Figure(figsize=(10, 4.2), dpi=96)
        self.price_fig.patch.set_facecolor(CARD)
        # 2 субплота: 75% цена, 25% объём
        self.price_ax  = self.price_fig.add_axes([0.07, 0.30, 0.90, 0.63])
        self.vol_ax    = self.price_fig.add_axes([0.07, 0.05, 0.90, 0.22],
                                                  sharex=self.price_ax)
        self.price_canvas = FigureCanvasTkAgg(self.price_fig, master=pc)
        self.price_canvas.get_tk_widget().grid(
            row=1, padx=8, pady=(0,10), sticky="nsew")
        self._draw_empty(self.price_ax, "Загрузи тикер на странице «Оценка»")
        self.price_canvas.draw()

        # Интерактивность — движение мыши
        self.price_canvas.mpl_connect("motion_notify_event", self._on_price_hover)
        self.price_canvas.mpl_connect("axes_leave_event",    self._on_price_leave)
        self.price_canvas.mpl_connect("button_press_event",  self._on_price_click)

        # ── График дивидендов ──────────────────────────────────
        dc = Card(self)
        dc.grid(row=2, column=0, padx=16, pady=(0,16), sticky="nsew")
        dc.grid_columnconfigure(0, weight=1)
        dc.grid_rowconfigure(1, weight=1)

        div_hdr = ctk.CTkFrame(dc, fg_color="transparent")
        div_hdr.grid(row=0, padx=16, pady=(12,4), sticky="ew")
        lbl(div_hdr, "История дивидендов", size=13,
            weight="bold", color=GOLD).pack(side="left")
        self.div_info = lbl(div_hdr, "", size=12, color=MUTED)
        self.div_info.pack(side="right")

        self.div_fig = Figure(figsize=(10, 2.8), dpi=96)
        self.div_fig.patch.set_facecolor(CARD)
        self.div_ax  = self.div_fig.add_subplot(111)
        self.div_canvas = FigureCanvasTkAgg(self.div_fig, master=dc)
        self.div_canvas.get_tk_widget().grid(
            row=1, padx=8, pady=(0,10), sticky="nsew")
        self._draw_empty(self.div_ax, "Загрузи тикер на странице «Оценка»")
        self.div_canvas.draw()

        self.div_canvas.mpl_connect("motion_notify_event", self._on_div_hover)
        self.div_canvas.mpl_connect("axes_leave_event",    self._on_div_leave)

        # Аннотации (создаём заранее, скрываем)
        self._div_annot = None

    # ── Пустой график ──────────────────────────────────────────
    def _draw_empty(self, ax, msg=""):
        ax.clear()
        ax.set_facecolor(CARD2)
        if msg:
            ax.text(0.5, 0.5, msg, transform=ax.transAxes,
                    ha="center", va="center", color=MUTED, fontsize=12)
        ax.set_xticks([]); ax.set_yticks([])
        for sp in ax.spines.values():
            sp.set_color(BORDER)

    # ── Загрузка данных ────────────────────────────────────────
    def load_ticker(self, ticker: str):
        self.current_ticker = ticker
        self.title_lbl.configure(text=f"Аналитика — {ticker}")
        self._all_dates = []; self._all_closes = []
        self._all_opens = []; self._all_volumes = []
        threading.Thread(target=self._fetch, args=(ticker,), daemon=True).start()

    def _fetch_candles_all(self, ticker):
        """Грузим всю историю дневных свечей с пагинацией MOEX ISS.
        Возвращает (dates, closes, opens, volumes)."""
        dates, closes, opens, volumes = [], [], [], []
        start = "2015-01-01"
        till  = datetime.now().strftime("%Y-%m-%d")
        while True:
            url = (f"https://iss.moex.com/iss/engines/stock/markets/shares/"
                   f"boards/TQBR/securities/{ticker}/candles.json"
                   f"?from={start}&till={till}"
                   f"&interval=24&iss.meta=off&iss.json=extended&start={len(dates)}")
            r = requests.get(url, timeout=20,
                             headers={"User-Agent": "StockValuator/1.0"})
            r.raise_for_status()
            batch_dates, batch_closes, batch_opens, batch_volumes = [], [], [], []
            for block in r.json():
                if not isinstance(block, dict): continue
                for row in block.get("candles", []):
                    try:
                        dt = datetime.strptime(row["begin"][:10], "%Y-%m-%d")
                        batch_dates.append(dt)
                        batch_closes.append(float(row["close"]))
                        batch_opens.append(float(row.get("open", row["close"])))
                        batch_volumes.append(float(row.get("volume", 0)))
                    except Exception:
                        continue
            if not batch_dates:
                break
            dates.extend(batch_dates); closes.extend(batch_closes)
            opens.extend(batch_opens); volumes.extend(batch_volumes)
            if len(batch_dates) < 500:
                break
        return dates, closes, opens, volumes

    def _fetch(self, ticker):
        # ── 1. Цена (все дневные свечи) ────────────────────────
        try:
            dates, closes, opens, volumes = self._fetch_candles_all(ticker)
            # Фильтруем: берём только данные за последние 10 лет
            cutoff10 = datetime.now() - timedelta(days=3650)
            quads = [(d, c, o, v) for d, c, o, v
                     in zip(dates, closes, opens, volumes) if d >= cutoff10]
            if quads:
                dates, closes, opens, volumes = map(list, zip(*quads))
            self._all_dates   = dates
            self._all_closes  = closes
            self._all_opens   = opens
            self._all_volumes = volumes
            self.after(0, lambda: self._apply_period(self._active_period))
        except Exception as e:
            self._all_dates  = []
            self._all_closes = []
            self.after(0, lambda: (
                self._draw_empty(self.price_ax, f"Ошибка загрузки цены: {e}"),
                self.price_canvas.draw()))

        # ── 2. Дивиденды — рисуем ПОСЛЕ загрузки цен ──────────
        try:
            divs = moex_dividends(ticker)
            self._divs_cache = divs
            self.after(0, lambda: self._draw_divs(divs, ticker))
        except Exception as e:
            self.after(0, lambda: (
                self._draw_empty(self.div_ax, f"Ошибка дивидендов: {e}"),
                self.div_canvas.draw()))

    # ── Переключение периода ───────────────────────────────────
    def _set_period(self, p):
        self._active_period = p
        for k, b in self._period_btns.items():
            b.configure(fg_color=ACCENT if k==p else CARD2)
        self._apply_period(p)

    def _apply_period(self, p):
        if p == "1Д":
            threading.Thread(
                target=self._fetch_intraday,
                args=(self.current_ticker,), daemon=True).start()
            return
        if not self._all_dates:
            return
        days, _ = PERIODS[p]
        all_v = list(zip(self._all_dates, self._all_closes,
                         self._all_opens  or [None]*len(self._all_dates),
                         self._all_volumes or [0]*len(self._all_dates)))
        if days == 0:
            filtered = all_v
        else:
            cutoff = datetime.now() - timedelta(days=days)
            filtered = [(d, c, o, v) for d, c, o, v in all_v if d >= cutoff]
            if not filtered:
                filtered = all_v[-min(30, len(all_v)):]
        dates, closes, opens, volumes = map(list, zip(*filtered)) if filtered else ([], [], [], [])
        self._draw_price(dates, closes, opens, volumes)

    def _fetch_intraday(self, ticker):
        """Минутные/10-минутные свечи за последние 2 дня."""
        try:
            since = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d")
            till  = datetime.now().strftime("%Y-%m-%d")
            url = (f"https://iss.moex.com/iss/engines/stock/markets/shares/"
                   f"boards/TQBR/securities/{ticker}/candles.json"
                   f"?from={since}&till={till}"
                   f"&interval=10&iss.meta=off&iss.json=extended")
            r = requests.get(url, timeout=15,
                             headers={"User-Agent": "StockValuator/1.0"})
            r.raise_for_status()
            dates, closes = [], []
            for block in r.json():
                if not isinstance(block, dict): continue
                for row in block.get("candles", []):
                    try:
                        dt = datetime.strptime(row["begin"][:16], "%Y-%m-%d %H:%M")
                        dates.append(dt)
                        closes.append(float(row["close"]))
                    except Exception:
                        continue
            if dates:
                self.after(0, lambda: self._draw_price(dates, closes))
            else:
                # Fallback: последний день из дневных свечей
                self.after(0, lambda: self._apply_period("1Н"))
        except Exception as e:
            self.after(0, lambda: self._apply_period("1Н"))

    def _load_compare(self):
        ticker = self._cmp_entry.get().strip().upper().replace(".ME", "")
        if not ticker or not self.current_ticker:
            return
        self._cmp_ticker = ticker
        self._cmp_dates  = []
        self._cmp_closes = []
        def worker():
            try:
                dates, closes, _, _ = self._fetch_candles_all(ticker)
                cutoff10 = datetime.now() - timedelta(days=3650)
                pairs = [(d, c) for d, c in zip(dates, closes) if d >= cutoff10]
                if pairs:
                    ds, cs = map(list, zip(*pairs))
                else:
                    ds, cs = [], []
                self._cmp_dates  = ds
                self._cmp_closes = cs
                self.after(0, lambda: self._apply_period(self._active_period))
                self.after(0, lambda: self._cmp_clear_btn.configure(state="normal"))
            except Exception:
                self._cmp_dates = []
                self._cmp_closes = []
        threading.Thread(target=worker, daemon=True).start()

    def _clear_compare(self):
        self._cmp_ticker  = None
        self._cmp_dates   = []
        self._cmp_closes  = []
        self._cmp_entry.delete(0, "end")
        self._cmp_clear_btn.configure(state="disabled")
        self.price_ax.set_ylabel("Цена, ₽", color=MUTED, fontsize=10)
        self._apply_period(self._active_period)

    def _on_norm_toggle(self):
        self._norm_mode = self._norm_var.get()
        if self._cmp_ticker:
            self._apply_period(self._active_period)

    def _on_volume_toggle(self):
        self.vol_ax.set_visible(self._show_volume.get())
        self.price_canvas.draw_idle()

    def _on_indicator_toggle(self):
        CONFIG["indicators"] = {k: v.get() for k, v in self._ind_vars.items()}
        _save_config()
        if hasattr(self, "_price_dates") and self._price_dates:
            self._draw_price(
                self._price_dates, self._price_closes,
                self._price_opens or None,
                self._price_vols or None,
            )

    # ── Рисуем цену ────────────────────────────────────────────
    def _draw_price(self, dates, closes, opens=None, volumes=None):
        ax = self.price_ax
        ax.clear()
        ax.set_facecolor(CARD2)
        self._annot_price    = None
        self._vline          = None
        self._pin_idx        = None
        self._pin_marker     = None
        self._price_span_patch = None

        if not dates:
            self._draw_empty(ax, "Нет данных")
            self.price_canvas.draw()
            return

        # ── Сравнение тикеров ──────────────────────────────────
        has_cmp = bool(self._cmp_dates and self._cmp_closes and self._cmp_ticker)
        if has_cmp:
            # Пересечение дат
            cmp_map = dict(zip(self._cmp_dates, self._cmp_closes))
            common_dates = sorted(d for d in dates if d in cmp_map)
            if common_dates:
                base_date = common_dates[0]
                idx0_main = dates.index(base_date)
                base_main = closes[idx0_main]
                base_cmp  = cmp_map[base_date]
                if self._norm_mode and base_main and base_cmp:
                    plot_main = [closes[dates.index(d)] / base_main * 100
                                 for d in common_dates]
                    plot_cmp  = [cmp_map[d] / base_cmp * 100
                                 for d in common_dates]
                    ax.set_ylabel("Норм. цена, %", color=MUTED, fontsize=10)
                else:
                    plot_main = [closes[dates.index(d)] for d in common_dates]
                    plot_cmp  = [cmp_map[d] for d in common_dates]
                ax.plot(common_dates, plot_main, color=GREEN,  linewidth=1.8,
                        label=self.current_ticker)
                ax.plot(common_dates, plot_cmp,  color="#79C0FF", linewidth=1.6,
                        label=self._cmp_ticker, alpha=0.9)
                ax.legend(fontsize=9, loc="upper left",
                          facecolor=CARD, edgecolor=BORDER, labelcolor=TEXT)
                dates  = common_dates
                closes = plot_main
            else:
                has_cmp = False

        if not has_cmp:
            ax.plot(dates, closes, color=GREEN, linewidth=1.8)
        self._price_span_patch = None

        # min / max маркеры
        mn, mx = min(closes), max(closes)
        mni, mxi = closes.index(mn), closes.index(mx)
        ax.scatter([dates[mni]], [mn], color=RED,   s=40, zorder=5)
        ax.scatter([dates[mxi]], [mx], color=GREEN, s=40, zorder=5)
        ax.annotate(f"мин {mn:.0f}", xy=(dates[mni], mn),
            xytext=(0,-20), textcoords="offset points",
            color=RED, fontsize=8, ha="center",
            arrowprops=dict(arrowstyle="-", color=RED, alpha=0.4, lw=0.8))
        ax.annotate(f"макс {mx:.0f}", xy=(dates[mxi], mx),
            xytext=(0,12), textcoords="offset points",
            color=GREEN, fontsize=8, ha="center",
            arrowprops=dict(arrowstyle="-", color=GREEN, alpha=0.4, lw=0.8))

        # Форматирование осей — автоматическое прореживание
        span = (dates[-1] - dates[0]).days
        if span <= 2:
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%d %b %H:%M"))
            ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
        elif span <= 14:
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%d %b"))
            ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
        elif span <= 90:
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%d %b"))
            ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=2))
        elif span <= 365:
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
            ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
        elif span <= 365 * 3:
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%b '%y"))
            ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
        else:
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
            ax.xaxis.set_major_locator(mdates.YearLocator())

        self.price_fig.autofmt_xdate(rotation=30, ha="right")
        ax.set_ylabel("Цена, ₽", color=MUTED, fontsize=10)
        ax.grid(True, axis="y", alpha=0.3, linestyle="--")
        for sp in ax.spines.values():
            sp.set_color(BORDER)

        # Аннотация (скрытая)
        self._annot_price = ax.annotate(
            "", xy=(dates[0], closes[0]),
            xytext=(12, 12), textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.5", fc=CARD, ec=ACCENT, lw=1),
            fontsize=10, color=TEXT, zorder=10)
        self._annot_price.set_visible(False)
        self._vline = ax.axvline(x=dates[0], color=ACCENT,
                                  alpha=0.5, linewidth=1, linestyle="--")
        self._vline.set_visible(False)

        self._price_dates  = dates
        self._price_closes = closes
        self._price_opens  = opens or []
        self._price_vols   = volumes or []

        # ── MA индикаторы ──────────────────────────────────────
        self._price_ma = {}
        _ma_styles = {
            "SMA20":  ("#FF9500", 1.2, "SMA 20"),
            "SMA50":  ("#9B59B6", 1.4, "SMA 50"),
            "SMA200": ("#F1C40F", 1.8, "SMA 200"),
            "EMA20":  ("#00BCD4", 1.2, "EMA 20"),
        }
        n = len(closes)
        for key, (ma_color, ma_lw, ma_label) in _ma_styles.items():
            if not self._ind_vars.get(key, tk.BooleanVar()).get():
                continue
            period = int(key.replace("SMA", "").replace("EMA", ""))
            if n < period:
                continue
            if key.startswith("SMA"):
                vals = [None] * (period - 1) + [
                    sum(closes[i - period:i]) / period for i in range(period, n + 1)
                ]
            else:  # EMA
                k_ema = 2 / (period + 1)
                ema_val = sum(closes[:period]) / period
                vals = [None] * (period - 1) + [ema_val]
                for i in range(period, n):
                    ema_val = closes[i] * k_ema + ema_val * (1 - k_ema)
                    vals.append(ema_val)
            self._price_ma[key] = vals
            plot_d = [dates[i] for i in range(n) if vals[i] is not None]
            plot_v = [vals[i]  for i in range(n) if vals[i] is not None]
            ax.plot(plot_d, plot_v, color=ma_color, linewidth=ma_lw,
                    label=ma_label, alpha=0.85)
        if self._price_ma:
            ax.legend(fontsize=8, loc="upper left",
                      facecolor=CARD, edgecolor=BORDER, labelcolor=TEXT)

        # ── Объём ──────────────────────────────────────────────
        vax = self.vol_ax
        vax.clear()
        vax.set_facecolor(CARD2)
        if volumes and opens and self._show_volume.get():
            bar_colors = [GREEN if c >= o else RED
                          for c, o in zip(closes, opens)]
            vax.bar(dates, volumes, color=bar_colors, alpha=0.7, width=1.0)
            vax.set_ylabel("Объём", color=MUTED, fontsize=8)
            vax.yaxis.set_major_formatter(
                plt.FuncFormatter(lambda x, _: f"{x/1e6:.0f}М" if x >= 1e6 else f"{x:.0f}"))
            vax.tick_params(axis="y", labelsize=7, colors=MUTED)
            vax.tick_params(axis="x", labelbottom=False)
            vax.grid(True, axis="y", alpha=0.2, linestyle="--")
            for sp in vax.spines.values():
                sp.set_color(BORDER)
        else:
            vax.set_visible(False)
        vax.set_visible(bool(volumes) and self._show_volume.get())

        # Изменение за период
        chg = (closes[-1]/closes[0] - 1)*100 if closes[0] else 0
        chg_color = GREEN if chg >= 0 else RED
        self.price_info.configure(
            text=f"{closes[-1]:.2f} ₽   {chg:+.1f}%", text_color=chg_color)

        self.price_canvas.draw()

    # ── Hover на графике цены ──────────────────────────────────
    def _on_price_hover(self, event):
        if (event.inaxes != self.price_ax
                or not hasattr(self, "_price_dates")
                or not self._price_dates
                or self._annot_price is None):
            return
        # Найти ближайшую точку
        import bisect
        dates  = self._price_dates
        closes = self._price_closes
        if not dates: return

        # Конвертируем x в datetime
        try:
            x_dt = mdates.num2date(event.xdata).replace(tzinfo=None)
        except Exception:
            return

        xs = [d for d in dates]
        idx = bisect.bisect_left(xs, x_dt)
        idx = max(0, min(idx, len(dates)-1))

        d  = dates[idx]
        cl = closes[idx]

        span = (self._price_dates[-1] - self._price_dates[0]).days
        fmt  = '%d %b %Y  %H:%M' if span <= 7 else '%d %b %Y'
        
        # Режим сравнения двух точек
        if self._pin_idx is not None:
            # Всегда от более ранней к более поздней дате
            if self._pin_idx <= idx:
                i_from, i_to = self._pin_idx, idx
            else:
                i_from, i_to = idx, self._pin_idx
            d_from  = self._price_dates[i_from]
            d_to    = self._price_dates[i_to]
            c_from  = self._price_closes[i_from]
            c_to    = self._price_closes[i_to]
            diff_rub = c_to - c_from
            diff_pct = (c_to / c_from - 1) * 100 if c_from else 0
            sign = '+' if diff_rub >= 0 else ''
            arrow = '▲' if diff_rub >= 0 else '▼'
            ec_color = GREEN if diff_rub >= 0 else RED
            self._annot_price.set_text(
                f"{d_from.strftime(fmt)}  →  {d_to.strftime(fmt)}\n"
                f"{c_from:.2f} ₽  →  {c_to:.2f} ₽\n"
                f"{arrow} {sign}{diff_rub:.2f} ₽  ({sign}{diff_pct:.1f}%)")
            self._annot_price.set(
                bbox=dict(boxstyle='round,pad=0.5', fc=CARD, ec=ec_color, lw=1.5))

            # Рисуем/обновляем зелёный коридор между двумя точками
            if self._price_span_patch:
                try: self._price_span_patch.remove()
                except: pass
            xmin = mdates.date2num(d_from)
            xmax = mdates.date2num(d_to)
            span_color = GREEN if diff_rub >= 0 else RED
            self._price_span_patch = self.price_ax.axvspan(
                xmin, xmax, alpha=0.15, color=span_color, zorder=1)
        else:
            # Убираем коридор если нет закреплённой точки
            if self._price_span_patch:
                try: self._price_span_patch.remove()
                except: pass
                self._price_span_patch = None
            vol_str = ""
            if hasattr(self, "_price_vols") and self._price_vols and idx < len(self._price_vols):
                v = self._price_vols[idx]
                vol_str = f"\nОбъём: {v/1e6:.2f}М" if v >= 1e6 else f"\nОбъём: {v:.0f}"
            ma_str = ""
            if hasattr(self, "_price_ma") and self._price_ma:
                ma_parts = []
                for _k in ("SMA20", "SMA50", "SMA200", "EMA20"):
                    if _k in self._price_ma and idx < len(self._price_ma[_k]):
                        _v = self._price_ma[_k][idx]
                        if _v is not None:
                            ma_parts.append(f"{_k}: {_v:.2f}")
                if ma_parts:
                    ma_str = "\n" + "  ".join(ma_parts)
            self._annot_price.set_text(f"{d.strftime(fmt)}\n{cl:.2f} ₽{vol_str}{ma_str}")
            self._annot_price.set(
                bbox=dict(boxstyle='round,pad=0.5', fc=CARD, ec=ACCENT, lw=1))

        self._annot_price.xy = (d, cl)
        self._annot_price.set_visible(True)
        self._vline.set_xdata([d])
        self._vline.set_visible(True)

        # Динамическое позиционирование: если курсор в правой половине →
        # тултип рисуем слева от точки, чтобы не выходил за край
        ax = self.price_ax
        xlim = ax.get_xlim()
        x_frac = (event.xdata - xlim[0]) / (xlim[1] - xlim[0]) if xlim[1] != xlim[0] else 0.5
        ylim = ax.get_ylim()
        y_frac = (cl - ylim[0]) / (ylim[1] - ylim[0]) if ylim[1] != ylim[0] else 0.5
        
        # Горизонталь: правее 55% → рисуем влево
        x_off = -14 if x_frac > 0.55 else 14
        # Вертикаль: выше 70% → рисуем вниз
        y_off = -14 if y_frac > 0.70 else 14
        
        ha = 'right' if x_frac > 0.55 else 'left'
        self._annot_price.set_ha(ha)
        self._annot_price.xyann = (x_off, y_off)
        
        # Маркер закреплённой точки
        if self._pin_marker:
            try: self._pin_marker.remove()
            except: pass
        if self._pin_idx is not None:
            pd_ = self._price_dates[self._pin_idx]
            pc_ = self._price_closes[self._pin_idx]
            self._pin_marker = self.price_ax.scatter(
                [pd_], [pc_], color=GOLD, s=60, zorder=11)
        
        self.price_canvas.draw_idle()

    def _on_price_leave(self, event):
        if self._pin_idx is None:   # только если нет закреплённой точки
            if self._annot_price:
                self._annot_price.set_visible(False)
            if self._vline:
                self._vline.set_visible(False)
            self.price_canvas.draw_idle()

    def _on_price_click(self, event):
        if event.inaxes != self.price_ax or not hasattr(self, '_price_dates'):
            return
        import bisect
        try:
            x_dt = mdates.num2date(event.xdata).replace(tzinfo=None)
        except Exception:
            return
        idx = bisect.bisect_left(self._price_dates, x_dt)
        idx = max(0, min(idx, len(self._price_dates)-1))

        if self._pin_idx is None:
            # Первый клик — закрепляем точку
            self._pin_idx = idx
        else:
            # Второй клик — сбрасываем всё
            self._pin_idx = None
            if self._pin_marker:
                try: self._pin_marker.remove()
                except: pass
                self._pin_marker = None
            if self._price_span_patch:
                try: self._price_span_patch.remove()
                except: pass
                self._price_span_patch = None
            if self._annot_price:
                self._annot_price.set_visible(False)
            if self._vline:
                self._vline.set_visible(False)
        self.price_canvas.draw_idle()

    # ── Рисуем дивиденды ───────────────────────────────────────
    def _draw_divs(self, divs, ticker):
        ax = self.div_ax
        ax.clear()
        ax.set_facecolor(CARD2)
        self._div_annot    = None
        self._div_bars_data = []

        if not divs:
            self._draw_empty(ax, "История дивидендов не найдена")
            self.div_canvas.draw()
            return

        # Группируем по году (суммируем несколько выплат в год)
        from collections import defaultdict
        year_divs = defaultdict(float)
        year_raw  = {}
        for raw_d, val in divs:
            yr = raw_d[:4]
            year_divs[yr] += val
            year_raw[yr]   = raw_d
        divs_by_year = sorted(year_divs.items())
        years     = [yr for yr, _ in divs_by_year]
        values    = [val for _, val in divs_by_year]
        raw_dates = [year_raw[yr] for yr in years]

        # Доходность: дивиденд / средняя цена за год
        yields = []
        for yr, val in zip(years, values):
            try:
                prices_year = [c for dt, c in zip(
                    getattr(self, "_all_dates",  []),
                    getattr(self, "_all_closes", []))
                    if dt.strftime("%Y") == yr]
                avg_p = sum(prices_year) / len(prices_year) if prices_year else 0
                yields.append(round(val / avg_p * 100, 1) if avg_p else 0)
            except Exception:
                yields.append(0)

        n = len(values)
        x = np.arange(n)

        ax.bar(x, values, width=0.6, zorder=3,
               color=GREEN_D, edgecolor=GREEN, linewidth=1.0)

        # Подписи суммы над столбцами
        for i, val in enumerate(values):
            ax.text(i, val + max(values) * 0.03,
                    f"{val:.2f} ₽",
                    ha="center", va="bottom",
                    fontsize=8, color=GREEN,
                    fontweight="bold", zorder=5)

        ax.set_xticks(x)
        ax.set_xticklabels(years, rotation=30, ha="right", fontsize=9)
        ax.set_ylabel("Дивиденд, ₽", color=MUTED, fontsize=10)
        ax.grid(True, axis="y", alpha=0.25, linestyle="--")
        for sp in ax.spines.values():
            sp.set_color(BORDER)

        # Аннотация hover
        self._div_annot = ax.annotate(
            "", xy=(0, 0), xytext=(0, 28), textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.5", fc=CARD, ec=GREEN, lw=1.5),
            fontsize=10, color=TEXT, zorder=10, ha="center")
        self._div_annot.set_visible(False)

        self._div_bars_data = list(zip(x, values, years, yields))

        # Итог: последний год
        last_yr  = years[-1] if years else ""
        last_val = values[-1] if values else 0
        self.div_info.configure(
            text=f"за {last_yr}: {last_val:.2f} ₽", text_color=GREEN)

        self.div_fig.subplots_adjust(top=0.88, bottom=0.22, left=0.09, right=0.97)
        self.div_canvas.draw()


    def _gold_shade(self, t):
        """t от 0 до 1 → от тёмного к золотому"""
        r = int(0x2D + t * (0xE3 - 0x2D))
        g = int(0x20 + t * (0xB3 - 0x20))
        b = int(0x08 + t * (0x41 - 0x08))
        return f"#{r:02x}{g:02x}{b:02x}"

    def _on_div_hover(self, event):
        if event.inaxes != self.div_ax or self._div_annot is None:
            return
        for xi, val, date_s, yld in self._div_bars_data:
            if abs(event.xdata - xi) < 0.4:
                txt = f"{date_s}\n{val:.2f} ₽"
                if yld > 0:
                    txt += f"\nДоходность: {yld:.1f}%"
                self._div_annot.set_text(txt)
                self._div_annot.xy = (xi, val)
                self._div_annot.set_visible(True)
                # Позиционирование: правее 60% → тултип влево, иначе вправо
                ax = self.div_ax
                xlim = ax.get_xlim()
                ylim = ax.get_ylim()
                x_frac = (xi - xlim[0]) / (xlim[1] - xlim[0]) if xlim[1] != xlim[0] else 0.5
                y_frac = (val - ylim[0]) / (ylim[1] - ylim[0]) if ylim[1] != ylim[0] else 0.5
                # Горизонталь
                if x_frac > 0.75:
                    x_off, ha = -60, 'center'
                elif x_frac < 0.25:
                    x_off, ha = 60, 'center'
                else:
                    x_off, ha = 0, 'center'
                # Вертикаль: если столбец высокий → тултип сбоку внизу
                y_off = 28 if y_frac < 0.75 else -45
                self._div_annot.set_ha(ha)
                self._div_annot.xyann = (x_off, y_off)
                self.div_canvas.draw_idle()
                return
        self._div_annot.set_visible(False)
        self.div_canvas.draw_idle()

    def _on_div_leave(self, event):
        if self._div_annot:
            self._div_annot.set_visible(False)
            self.div_canvas.draw_idle()


# ═══════════════════════════════════════════════════════════════
#  Страница 3 — Настройки
# ═══════════════════════════════════════════════════════════════
class SettingsPage(ctk.CTkFrame):
    def __init__(self, parent, app, **kw):
        super().__init__(parent, fg_color=BG, **kw)
        self.app = app
        self._build()

    def _build(self):
        wrap = ctk.CTkFrame(self, fg_color="transparent")
        wrap.pack(padx=40, pady=40, anchor="nw")

        lbl(wrap, "Настройки", size=22, weight="bold").pack(anchor="w")
        lbl(wrap, "Параметры ставок дисконтирования",
            size=13, color=MUTED).pack(anchor="w", pady=(4, 24))

        # ── Карточка с полями ───────────────────────────────────
        card = Card(wrap)
        card.pack(anchor="w")
        card.grid_columnconfigure(1, weight=1)

        fields_cfg = [
            ("r_f", "Безрисковая ставка  r_f",
             "Ключевая ставка ЦБ РФ",
             "0.1600"),
            ("r_m", "Доходность рынка  r_m",
             "5-летний CAGR индекса MCFTR (полная доходность MOEX)",
             "0.2200"),
        ]

        self._entries  = {}
        self._src_lbls = {}   # подписи-источники рядом с полями
        for ri, (key, title, hint, default) in enumerate(fields_cfg):
            if ri > 0:
                Div(card).grid(row=ri*3 - 1, column=0, columnspan=3,
                               padx=20, pady=0, sticky="ew")

            lbl(card, title, size=13, weight="bold").grid(
                row=ri*3, column=0, padx=(20, 40), pady=(20, 2), sticky="w")
            lbl(card, hint, size=11, color=MUTED).grid(
                row=ri*3+1, column=0, padx=(20, 40), pady=(0, 16), sticky="w")

            e = inp(card, width=160, ph=default)
            e.insert(0, str(CONFIG.get(key, default)))
            e.grid(row=ri*3, column=1, rowspan=2,
                   padx=(0, 12), pady=16, sticky="w")
            self._entries[key] = e

            # Метка-источник (заполняется после авто-загрузки)
            sl = lbl(card, "", size=10, color=MUTED)
            sl.grid(row=ri*3, column=2, rowspan=2,
                    padx=(0, 20), pady=16, sticky="w")
            self._src_lbls[key] = sl

        # ── Кнопки ─────────────────────────────────────────────
        row_btn = ctk.CTkFrame(wrap, fg_color="transparent")
        row_btn.pack(anchor="w", pady=(16, 0))

        self._auto_btn = btn(row_btn, "⟳ Загрузить из интернета",
                             self._auto_fetch,
                             color=ACCENT, width=220, height=38)
        self._auto_btn.pack(side="left")
        btn(row_btn, "Применить", self._apply,
            color="#238636", width=140, height=38).pack(side="left", padx=(10, 0))
        btn(row_btn, "Сбросить", self._reset,
            color=CARD2, width=120, height=38).pack(side="left", padx=(10, 0))
        self._status_lbl = lbl(row_btn, "", size=12, color=MUTED)
        self._status_lbl.pack(side="left", padx=(16, 0))

        # ── Отраслевые P/E ──────────────────────────────────────
        lbl(wrap, "Отраслевые P/E", size=16, weight="bold").pack(
            anchor="w", pady=(32, 4))
        lbl(wrap, "Используются в модели сравнений (P × EPS). "
            "Загружаются автоматически при старте приложения.",
            size=12, color=MUTED).pack(anchor="w", pady=(0, 12))

        pe_card = Card(wrap)
        pe_card.pack(anchor="w", fill="x")

        # Заголовки таблицы
        hdr_f = ctk.CTkFrame(pe_card, fg_color="transparent")
        hdr_f.pack(fill="x", padx=20, pady=(14, 4))
        for txt, w in [("Сектор", 180), ("Базовый", 80), ("Актуальный", 100)]:
            lbl(hdr_f, txt, size=11, color=MUTED, weight="bold", width=w).pack(
                side="left", anchor="w")

        Div(pe_card).pack(fill="x", padx=20, pady=2)

        self._pe_rows: dict = {}   # sector → (base_lbl, live_lbl)
        sectors_order = [
            "Банки", "Нефть и газ", "Металлы", "Ритейл",
            "Телеком", "Технологии", "Электроэнергетика",
            "Удобрения", "Транспорт", "Девелопмент",
        ]
        for sector in sectors_order:
            row_f = ctk.CTkFrame(pe_card, fg_color="transparent")
            row_f.pack(fill="x", padx=20, pady=2)
            lbl(row_f, sector, size=12, width=180).pack(side="left", anchor="w")
            base_val = SECTOR_PE.get(sector, SECTOR_PE["default"])
            base_l = lbl(row_f, f"{base_val:.1f}x", size=12, color=MUTED, width=80)
            base_l.pack(side="left", anchor="w")
            live_l = lbl(row_f, "—", size=12, color=MUTED)
            live_l.pack(side="left", anchor="w")
            self._pe_rows[sector] = (base_l, live_l)

        pe_btn_row = ctk.CTkFrame(pe_card, fg_color="transparent")
        pe_btn_row.pack(anchor="w", padx=20, pady=(10, 14))
        self._pe_btn = btn(pe_btn_row, "⟳ Обновить P/E секторов",
                           self._refresh_pe_clicked,
                           color=CARD2, width=200, height=32)
        self._pe_btn.pack(side="left")
        self._pe_status = lbl(pe_btn_row, "", size=11, color=MUTED)
        self._pe_status.pack(side="left", padx=(12, 0))

    def _auto_fetch(self):
        """Подтягивает ставки из интернета в фоне."""
        self._auto_btn.configure(state="disabled", text="Загружаю…")
        self._status_lbl.configure(text="", text_color=MUTED)

        def worker():
            results = {}
            # 1. ЦБ РФ ключевая ставка
            date_s, r_f = fetch_cbr_key_rate()
            if r_f is not None:
                results["r_f"] = (r_f, f"ЦБ РФ от {date_s}")
            # 2. MOEX историческая доходность (5 лет)
            cagr, desc = fetch_moex_market_return(years=5)
            if cagr is not None:
                results["r_m"] = (cagr, desc)
            self.after(0, lambda: self._fill_auto(results))

        threading.Thread(target=worker, daemon=True).start()

    def _fill_auto(self, results):
        self._auto_btn.configure(state="normal", text="⟳ Загрузить из интернета")
        if not results:
            self._status_lbl.configure(
                text="Не удалось загрузить — проверь интернет", text_color=RED)
            return
        for key, (val, src) in results.items():
            e = self._entries[key]
            e.delete(0, "end")
            e.insert(0, f"{val:.4f}")
            self._src_lbls[key].configure(text=src, text_color=GREEN)
        loaded = ", ".join(
            f"{k}={v*100:.2f}%" for k, (v, _) in results.items())
        self._status_lbl.configure(
            text=f"✓ Загружено: {loaded}  — нажми «Применить»",
            text_color=GREEN)

    def _apply(self):
        try:
            r_f = float(self._entries["r_f"].get().strip())
            r_m = float(self._entries["r_m"].get().strip())
        except ValueError:
            self._status_lbl.configure(
                text="Ошибка: введи числа (например 0.16)", text_color=RED)
            return
        if not (0 < r_f < 1):
            self._status_lbl.configure(
                text="r_f должна быть от 0 до 1 (ставка ЦБ не бывает отрицательной)", text_color=RED)
            return
        if not (-1 < r_m < 10):
            self._status_lbl.configure(
                text="r_m вне разумного диапазона (от -100% до +1000%)", text_color=RED)
            return
        CONFIG["r_f"] = r_f
        CONFIG["r_m"] = r_m
        _save_config()
        self._status_lbl.configure(
            text=f"✓ Сохранено: r_f={r_f*100:.2f}%  r_m={r_m*100:.2f}%",
            text_color=GREEN)
        self.app._refresh_rates_label()

    def _reset(self):
        CONFIG.update(_CONFIG_DEFAULTS)
        _save_config()
        for key, e in self._entries.items():
            e.delete(0, "end")
            e.insert(0, str(_CONFIG_DEFAULTS[key]))
        self._status_lbl.configure(
            text=f"↩ Сброшено: r_f={_CONFIG_DEFAULTS['r_f']*100:.0f}%  "
                 f"r_m={_CONFIG_DEFAULTS['r_m']*100:.0f}%",
            text_color=MUTED)
        self.app._refresh_rates_label()

    def _refresh_pe_clicked(self):
        self._pe_btn.configure(state="disabled", text="Загружаю…")
        self._pe_status.configure(text="Запрашиваю данные с MOEX и SmartLab…", text_color=MUTED)
        threading.Thread(target=self._do_refresh_pe, daemon=True).start()

    def _do_refresh_pe(self):
        data = fetch_sector_pe_live()
        self.after(0, lambda: self._fill_pe(data))

    def _fill_pe(self, data: dict):
        sv._LIVE_SECTOR_PE.update(data)
        for sector, (_, live_l) in self._pe_rows.items():
            val = data.get(sector)
            if val:
                live_l.configure(text=f"{val:.1f}x", text_color=GREEN)
            else:
                live_l.configure(text="—", text_color=MUTED)
        self._pe_btn.configure(state="normal", text="⟳ Обновить P/E секторов")
        now = datetime.now().strftime("%H:%M")
        self._pe_status.configure(
            text=f"✓ Обновлено в {now}", text_color=GREEN)


# ═══════════════════════════════════════════════════════════════
#  Страница 4 — Скрининг
# ═══════════════════════════════════════════════════════════════
class ScreeningPage(ctk.CTkFrame):
    _COLS = [
        ("ticker",  "Тикер",           80),
        ("name",    "Название",        180),
        ("price",   "Цена, ₽",         90),
        ("fair",    "Справ. цена, ₽",  110),
        ("upside",  "Upside, %",        90),
    ]

    def __init__(self, parent, app, **kw):
        super().__init__(parent, fg_color=BG, **kw)
        self.app = app
        self._results    = []   # список dict с результатами
        self._sort_col   = "upside"
        self._sort_rev   = True
        self._total      = 0
        self._done       = 0
        self._build()

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # Заголовок
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.grid(row=0, padx=16, pady=(14, 6), sticky="ew")
        lbl(hdr, "Скрининг акций", size=15, weight="bold").pack(side="left")

        # Панель ввода
        inp_card = Card(self)
        inp_card.grid(row=1, padx=16, pady=(0, 6), sticky="ew")
        inp_card.grid_columnconfigure(0, weight=1)

        top_row = ctk.CTkFrame(inp_card, fg_color="transparent")
        top_row.grid(row=0, padx=16, pady=(12, 6), sticky="ew")
        top_row.grid_columnconfigure(0, weight=1)

        lbl(top_row, "Тикеры (через запятую):", size=11,
            color=MUTED).grid(row=0, column=0, sticky="w")

        self._ticker_entry = ctk.CTkEntry(
            top_row, placeholder_text="SBER, LKOH, GAZP, NVTK, GMKN…",
            height=34, font=ctk.CTkFont(size=12))
        self._ticker_entry.grid(row=1, column=0, sticky="ew", padx=(0, 8))
        self._ticker_entry.bind("<Return>", lambda e: self._start_screen())

        btn_col = ctk.CTkFrame(top_row, fg_color="transparent")
        btn_col.grid(row=1, column=1)
        btn(btn_col, "Скринить", self._start_screen,
            color=ACCENT, width=100, height=34).pack(side="left", padx=(0, 6))
        btn(btn_col, "Импорт из портфеля", self._import_portfolio,
            color=CARD2, width=160, height=34).pack(side="left")

        # Прогресс
        prog_row = ctk.CTkFrame(inp_card, fg_color="transparent")
        prog_row.grid(row=1, padx=16, pady=(0, 12), sticky="ew")
        prog_row.grid_columnconfigure(0, weight=1)
        self._prog_bar = ctk.CTkProgressBar(prog_row, height=6)
        self._prog_bar.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        self._prog_bar.set(0)
        self._prog_lbl = lbl(prog_row, "", size=11, color=MUTED)
        self._prog_lbl.grid(row=0, column=1)

        # Таблица
        import tkinter.ttk as ttk
        tbl_card = Card(self)
        tbl_card.grid(row=2, padx=16, pady=(0, 16), sticky="nsew")
        tbl_card.grid_columnconfigure(0, weight=1)
        tbl_card.grid_rowconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Screen.Treeview",
                         background=CARD2, foreground=TEXT,
                         fieldbackground=CARD2, rowheight=30,
                         font=("SF Pro Display", 12))
        style.configure("Screen.Treeview.Heading",
                         background=CARD, foreground=MUTED,
                         font=("SF Pro Display", 11, "bold"), relief="flat")
        style.map("Screen.Treeview",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", TEXT)])

        self._tree = ttk.Treeview(
            tbl_card, style="Screen.Treeview",
            columns=[c[0] for c in self._COLS],
            show="headings", selectmode="browse")

        for col_id, col_name, col_w in self._COLS:
            self._tree.heading(col_id, text=col_name,
                               command=lambda c=col_id: self._sort_by(c))
            self._tree.column(col_id, width=col_w, anchor="center",
                              minwidth=60, stretch=True)
        self._tree.column("name", anchor="w")

        sb = ttk.Scrollbar(tbl_card, orient="vertical",
                           command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        tbl_card.grid_columnconfigure(0, weight=1)

        self._tree.bind("<Double-1>", self._on_row_click)
        self._tree.tag_configure("green", foreground=GREEN)
        self._tree.tag_configure("red",   foreground=RED)

    def _import_portfolio(self):
        portfolio = self.app.valuation_page.portfolio
        if not portfolio:
            return
        tickers = ", ".join(portfolio.keys())
        self._ticker_entry.delete(0, "end")
        self._ticker_entry.insert(0, tickers)

    def _start_screen(self):
        raw = self._ticker_entry.get().strip()
        if not raw:
            return
        tickers = [t.strip().upper() for t in raw.split(",") if t.strip()]
        if not tickers:
            return

        # Очищаем таблицу
        for row in self._tree.get_children():
            self._tree.delete(row)
        self._results = []
        self._total   = len(tickers)
        self._done    = 0
        self._prog_bar.set(0)
        self._prog_lbl.configure(text=f"0 / {self._total}")

        from concurrent.futures import ThreadPoolExecutor
        def run():
            with ThreadPoolExecutor(max_workers=5) as ex:
                for ticker in tickers:
                    ex.submit(self._fetch_one, ticker)
        threading.Thread(target=run, daemon=True).start()

    def _fetch_one(self, ticker):
        try:
            price = moex_price(ticker)
            name  = moex_name(ticker)
            sl    = fetch_smartlab(ticker)
            divs  = moex_dividends(ticker)
            d0 = 0.0
            if divs:
                cutoff = (datetime.now().replace(
                    year=datetime.now().year - 1)).strftime("%Y-%m-%d")
                recent = [v for dt, v in divs if dt >= cutoff]
                d0 = sum(recent) if recent else sum(v for _, v in divs[-4:])

            r_f = CONFIG["r_f"]; r_m = CONFIG["r_m"]
            eps  = sl.get("eps", 0) or 0
            bvps = sl.get("bvps", 0) or 0
            roe  = sl.get("roe", 0.15) or 0.15
            g    = sl.get("g", 0.08) or 0.08
            beta = sl.get("beta", 1.0) or 1.0
            r_capm = r_f + beta * (r_m - r_f)
            d1 = d0 * (1 + g)
            r_ddm = (d1 / price + g) if price and d1 else r_capm
            r_ddm = max(min(r_ddm, 0.60), r_f)
            r_avg = (r_capm + r_ddm) / 2
            pv_ri = sum((roe - r_avg) * bvps / (1 + r_avg) ** t
                        for t in range(1, 6)) if r_avg else 0

            d = dict(ticker=ticker, name=name, price=price,
                     d0=d0, d1=d1, g=g, k=r_avg,
                     eps=eps, pe=(price / eps if eps > 0 else 0),
                     bvps=bvps, pv_ri=pv_ri, roe=roe,
                     beta=beta, r_f=r_f, r_m=r_m,
                     r_capm=r_capm, r_ddm=r_ddm, r_avg=r_avg,
                     currency="₽")
            models = [v for v in [ddm_price(d), pe_price(d),
                                  riv_price(d), dcf_price(d)] if v > 0]
            fair = sum(models) / len(models) if models else 0
            upside = (fair / price - 1) * 100 if price and fair else 0
            result = dict(ticker=ticker, name=name or ticker,
                          price=price, fair=fair, upside=upside)
        except Exception:
            result = dict(ticker=ticker, name="—", price=0, fair=0, upside=0)
        self.after(0, lambda r=result: self._add_result(r))

    def _add_result(self, r):
        self._results.append(r)
        self._done += 1
        self._prog_bar.set(self._done / self._total if self._total else 0)
        self._prog_lbl.configure(text=f"{self._done} / {self._total}")
        self._refresh_table()

    def _refresh_table(self):
        key = self._sort_col
        rev = self._sort_rev
        try:
            sorted_res = sorted(self._results,
                                key=lambda x: x.get(key, 0) or 0,
                                reverse=rev)
        except Exception:
            sorted_res = self._results
        for row in self._tree.get_children():
            self._tree.delete(row)
        for r in sorted_res:
            up = r.get("upside", 0)
            tag = "green" if up > 0 else ("red" if up < 0 else "")
            up_str = f"+{up:.1f}%" if up > 0 else f"{up:.1f}%"
            self._tree.insert("", "end", iid=r["ticker"], tags=(tag,), values=(
                r["ticker"],
                r["name"],
                f"{r['price']:.2f}" if r["price"] else "—",
                f"{r['fair']:.2f}"  if r["fair"]  else "—",
                up_str if r["fair"] else "—",
            ))

    def _sort_by(self, col):
        if self._sort_col == col:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col
            self._sort_rev = col in ("upside", "fair", "price")
        self._refresh_table()

    def _on_row_click(self, event):
        sel = self._tree.selection()
        if not sel:
            return
        ticker = sel[0]
        self.app.valuation_page.ticker_e.delete(0, "end")
        self.app.valuation_page.ticker_e.insert(0, ticker)
        self.app._show_page("valuation")
        self.app.valuation_page._load_data()


# ═══════════════════════════════════════════════════════════════
#  Главное окно
# ═══════════════════════════════════════════════════════════════
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Stock Valuation")
        self.geometry("1300x820")
        self.minsize(1100, 700)
        self.configure(fg_color=BG)
        self._build_nav()
        self._build_pages()
        self._show_page("valuation")

    def _build_nav(self):
        nav = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, height=48)
        nav.pack(side="top", fill="x")
        nav.pack_propagate(False)

        lbl(nav, "Stock Valuation", size=15,
            weight="bold", color=TEXT).pack(side="left", padx=20)

        self.nav_btns = {}
        bf = ctk.CTkFrame(nav, fg_color="transparent")
        bf.pack(side="left", padx=16)
        for key, name in [("valuation","Оценка"),
                           ("analytics","График & Дивиденды"),
                           ("screening","Скрининг"),
                           ("settings", "Настройки")]:
            b = ctk.CTkButton(
                bf, text=name, width=180, height=32,
                fg_color="transparent", hover_color=CARD2,
                text_color=MUTED, font=ctk.CTkFont(size=13),
                corner_radius=6,
                command=lambda k=key: self._show_page(k))
            b.pack(side="left", padx=3)
            self.nav_btns[key] = b

        # Индикатор ставок справа в nav
        rates_frame = ctk.CTkFrame(nav, fg_color=CARD2,
                                   corner_radius=8)
        rates_frame.pack(side="right", padx=16, pady=8)
        self._rates_lbl = lbl(
            rates_frame,
            f"r_f {CONFIG['r_f']*100:.1f}%   r_m {CONFIG['r_m']*100:.1f}%",
            size=12, color=MUTED)
        self._rates_lbl.pack(padx=12, pady=4)

    def _build_pages(self):
        c = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        c.pack(fill="both", expand=True)
        c.grid_columnconfigure(0, weight=1)
        c.grid_rowconfigure(0, weight=1)

        self.analytics_page  = AnalyticsPage(c)
        self.valuation_page  = ValuationPage(c, app=self)
        self.settings_page   = SettingsPage(c, app=self)
        self.screening_page  = ScreeningPage(c, app=self)
        for p in (self.valuation_page, self.analytics_page,
                  self.settings_page, self.screening_page):
            p.grid(row=0, column=0, sticky="nsew")
        # Автозагрузка отраслевых P/E при старте
        threading.Thread(target=self._refresh_sector_pe, daemon=True).start()

    def _refresh_sector_pe(self):
        """Фоновый поток: загружает актуальные отраслевые P/E и обновляет UI."""
        try:
            data = fetch_sector_pe_live()
            self.after(0, lambda: self.settings_page._fill_pe(data))
        except Exception:
            pass

    def _show_page(self, key):
        {"valuation": self.valuation_page,
         "analytics": self.analytics_page,
         "settings":  self.settings_page,
         "screening": self.screening_page}[key].tkraise()
        for k, b in self.nav_btns.items():
            b.configure(fg_color=CARD2 if k==key else "transparent",
                        text_color=TEXT  if k==key else MUTED)

    def _refresh_rates_label(self):
        self._rates_lbl.configure(
            text=f"r_f {CONFIG['r_f']*100:.1f}%   r_m {CONFIG['r_m']*100:.1f}%")


if __name__ == "__main__":
    App().mainloop()
